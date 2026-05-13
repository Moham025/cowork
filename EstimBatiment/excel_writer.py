# excel_writer.py
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def create_excel_table_for_block(ws, start_row, roman_numeral_main, header_title, items_data_list):
    """
    Creates a formatted table for a given block on the specified worksheet,
    starting at 'start_row'.
    
    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): The OpenPyXL worksheet to write to.
        start_row (int): The row number to start writing this block from.
        roman_numeral_main (str): The Roman numeral for the block (e.g., "I", "II").
        header_title (str): The main title of the block (e.g., "TERRASSEMENT").
        items_data_list (list): List of lists/tuples, each sub-list containing
                                [description, unit, calculated_quantity, unit_price].
    
    Returns:
        tuple: A tuple containing (next_available_row, total_cell_reference, numeric_block_total).
               next_available_row: The number of the next available row after the end of this block.
               total_cell_reference: The Excel cell coordinate (e.g., "F9") where the block's total is located.
               numeric_block_total: The sum of (quantity * unit_price) for all items in this block.
    """
    num_data_rows = len(items_data_list)
    
    # Style definitions
    thin_black_side = Side(style='thin', color='000000')
    black_border = Border(left=thin_black_side, right=thin_black_side, top=thin_black_side, bottom=thin_black_side)
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Column width definitions
    column_widths = [6, 60, 6, 12, 10, 15] # A: Num, B: Desc, C: Unité, D: Qté, E: P.U., F: Montant
    column_letters = [get_column_letter(i + 1) for i in range(len(column_widths))]
    for i, width in enumerate(column_widths):
        ws.column_dimensions[column_letters[i]].width = width

    # Write main block header (Roman numeral and Title)
    ws[f'A{start_row}'] = roman_numeral_main
    ws[f'A{start_row}'].font = bold_font
    ws[f'A{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{start_row}'].fill = header_fill

    # Merge cells for block title
    ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=len(column_widths))
    ws[f'B{start_row}'] = header_title
    ws[f'B{start_row}'].font = bold_font
    ws[f'B{start_row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws[f'B{start_row}'].fill = header_fill

    # Initialize numeric total for this block
    numeric_block_total = 0.0

    # Write item data
    current_row_idx = start_row + 1 # Start data writing after the header
    for i, item_data in enumerate(items_data_list):
        description, unit, qty_calculated, unit_price = item_data
        
        ws[f'A{current_row_idx}'] = f"{roman_numeral_main}.{i + 1}"
        ws[f'A{current_row_idx}'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws[f'B{current_row_idx}'] = description
        ws[f'B{current_row_idx}'].alignment = Alignment(wrap_text=True, vertical='top')
        
        ws[f'C{current_row_idx}'] = unit
        ws[f'C{current_row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws[f'D{current_row_idx}'] = qty_calculated
        ws[f'D{current_row_idx}'].number_format = '#,##0.00'
        ws[f'D{current_row_idx}'].alignment = Alignment(horizontal='right', vertical='center')
        
        ws[f'E{current_row_idx}'] = unit_price
        ws[f'E{current_row_idx}'].number_format = '#,##0.00'
        ws[f'E{current_row_idx}'].alignment = Alignment(horizontal='right', vertical='center')
        
        # Formula for the amount (Qté * P.U.)
        item_amount = (qty_calculated if qty_calculated is not None else 0.0) * (unit_price if unit_price is not None else 0.0)
        ws[f'F{current_row_idx}'] = f"=D{current_row_idx}*E{current_row_idx}" # Excel formula
        ws[f'F{current_row_idx}'].number_format = '#,##0.00'
        ws[f'F{current_row_idx}'].alignment = Alignment(horizontal='right', vertical='center')
        
        numeric_block_total += item_amount # Accumulate numeric total
        current_row_idx += 1

    # Write TOTAL row for the block
    total_row_idx = current_row_idx
    ws.merge_cells(start_row=total_row_idx, start_column=1, end_row=total_row_idx, end_column=4)
    ws[f'A{total_row_idx}'] = f"TOTAL {roman_numeral_main}"
    ws[f'A{total_row_idx}'].font = bold_font
    ws[f'A{total_row_idx}'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws[f'E{total_row_idx}'] = "Somme"
    ws[f'E{total_row_idx}'].font = bold_font
    ws[f'E{total_row_idx}'].alignment = Alignment(horizontal='left', vertical='center')
    
    # Formula for the total sum of the block
    total_cell_ref = f"F{total_row_idx}" # Capture the cell reference for the total
    if num_data_rows > 0:
        sum_formula = f"=SUM(F{start_row + 1}:F{total_row_idx - 1})"
    else:
        sum_formula = 0
    ws[total_cell_ref] = sum_formula
    ws[total_cell_ref].font = bold_font
    ws[total_cell_ref].number_format = '#,##0.00'
    ws[total_cell_ref].alignment = Alignment(horizontal='right', vertical='center')

    # Apply borders to all cells in the block
    for row_num in range(start_row, total_row_idx + 1):
        for col_num in range(1, len(column_widths) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = black_border
            
    print(f"Bloc {roman_numeral_main} generated starting from row {start_row}.")
    # Return the next available row (with 2 blank rows for readability) AND the total cell reference AND the numeric total
    return total_row_idx + 2, total_cell_ref, numeric_block_total


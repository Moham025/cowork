# expression_engine.py
"""
Évaluateur d'expressions personnalisées pour EstimType2.xlsx.

Syntaxes supportées :
  NOM(col){feuille}   → INDEX/MATCH double : ligne où A=NOM, col où ligne1=col
  XN{feuille}         → cellule directe colonne X, ligne N
  NOM[col]            → raccourci pour NOM(col){qt}
  opérateurs + - * /  → arithmétique standard
  constantes          → virgule française acceptée (0,16 → 0.16)
"""

import re
import openpyxl
from openpyxl.utils import column_index_from_string

# Pré-compilation des patterns
_PAT_NAMED  = re.compile(r'([A-Za-z_][A-Za-z0-9_]*)\(([A-Za-z0-9_]+)\)\{([^}]+)\}')
_PAT_CELL   = re.compile(r'([A-Za-z]+)(\d+)\{([^}]+)\}')
_PAT_SHORT  = re.compile(r'([A-Za-z_][A-Za-z0-9_]*)\[([A-Za-z0-9_]+)\]')


class ExpressionEngine:
    """Moteur d'évaluation chargé une fois par fichier."""

    def __init__(self, wb_vals):
        """
        wb_vals : workbook openpyxl chargé avec data_only=True
                  (valeurs pré-calculées pour qt, cf, ha…)
        """
        self._wb = wb_vals
        self._sheet_cache: dict = {}    # name → worksheet
        self._lookup_cache: dict = {}   # (element, col, sheet) → float
        self._cell_cache: dict = {}     # (col_str, row, sheet) → float

    # ── Accès aux feuilles ────────────────────────────────────────────────────

    def _ws(self, name: str):
        if name not in self._sheet_cache:
            name_stripped = name.strip()
            if name_stripped in self._wb.sheetnames:
                self._sheet_cache[name] = self._wb[name_stripped]
            else:
                # Recherche insensible à la casse
                for sn in self._wb.sheetnames:
                    if sn.strip().lower() == name_stripped.lower():
                        self._sheet_cache[name] = self._wb[sn]
                        break
                else:
                    self._sheet_cache[name] = None
        return self._sheet_cache[name]

    # ── Lookup bidirectionnel NOM(col){sheet} ─────────────────────────────────

    def lookup_named(self, element: str, col_name: str, sheet_name: str) -> float:
        key = (element.lower(), col_name.lower(), sheet_name.strip().lower())
        if key in self._lookup_cache:
            return self._lookup_cache[key]

        ws = self._ws(sheet_name)
        if ws is None:
            self._lookup_cache[key] = 0.0
            return 0.0

        # Trouver la ligne où col A = element
        row_idx = None
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if v is not None and str(v).strip().lower() == element.lower():
                row_idx = r
                break

        if row_idx is None:
            self._lookup_cache[key] = 0.0
            return 0.0

        # Trouver la colonne où ligne 1 = col_name
        col_idx = None
        for c in range(1, ws.max_column + 1):
            v = ws.cell(1, c).value
            if v is not None and str(v).strip().lower() == col_name.lower():
                col_idx = c
                break

        if col_idx is None:
            self._lookup_cache[key] = 0.0
            return 0.0

        result = _to_float(ws.cell(row_idx, col_idx).value)
        self._lookup_cache[key] = result
        return result

    # ── Référence directe XN{sheet} ───────────────────────────────────────────

    def lookup_cell(self, col_letters: str, row_num: int, sheet_name: str) -> float:
        key = (col_letters.upper(), row_num, sheet_name.strip().lower())
        if key in self._cell_cache:
            return self._cell_cache[key]

        ws = self._ws(sheet_name)
        if ws is None:
            self._cell_cache[key] = 0.0
            return 0.0

        try:
            col_idx = column_index_from_string(col_letters)
            result = _to_float(ws.cell(row_num, col_idx).value)
        except Exception:
            result = 0.0

        self._cell_cache[key] = result
        return result

    # ── Évaluation d'expression ───────────────────────────────────────────────

    def evaluate(self, expr) -> float:
        """Parse et évalue une expression personnalisée."""
        if expr is None:
            return 0.0

        s = str(expr).strip()

        # Nombre direct (virgule FR ou point)
        try:
            return float(s.replace(',', '.'))
        except ValueError:
            pass

        # 1. Raccourci NOM[col] → NOM(col){qt}
        s = _PAT_SHORT.sub(lambda m: f'{m.group(1)}({m.group(2)}){{qt}}', s)

        # 2. NOM(col){sheet} → valeur numérique
        def _sub_named(m):
            val = self.lookup_named(m.group(1), m.group(2), m.group(3))
            return str(val)

        s = _PAT_NAMED.sub(_sub_named, s)

        # 3. XN{sheet} → valeur numérique
        def _sub_cell(m):
            val = self.lookup_cell(m.group(1), int(m.group(2)), m.group(3))
            return str(val)

        s = _PAT_CELL.sub(_sub_cell, s)

        # 4. Virgule française restante → point
        s = s.replace(',', '.')

        # 5. Évaluation arithmétique sécurisée
        try:
            result = eval(s, {"__builtins__": {}}, {})  # noqa: S307
            return float(result)
        except Exception:
            return 0.0


# ── Helpers ───────────────────────────────────────────────────────────────────

def _to_float(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).replace(',', '.').strip())
    except (ValueError, TypeError):
        return 0.0


def load_engine(file_bytes: bytes) -> tuple:
    """
    Charge le workbook et retourne (engine, wb_exprs).
    wb_exprs : chargé avec data_only=False pour lire les textes d'expression.
    engine   : ExpressionEngine sur wb chargé data_only=True (valeurs pré-calculées).
    """
    import io
    raw = io.BytesIO(file_bytes)
    wb_vals = openpyxl.load_workbook(raw, data_only=True)
    raw.seek(0)
    wb_exprs = openpyxl.load_workbook(raw, data_only=False)
    engine = ExpressionEngine(wb_vals)
    return engine, wb_exprs

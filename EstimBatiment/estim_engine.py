# estim_engine.py
"""
Moteur de traitement central pour EstimBatiment.
Gère les imports relatifs de calculation_engine via importlib.
"""
import sys
import os
import io
import types
import importlib.util
from datetime import datetime

_dir = os.path.dirname(os.path.abspath(__file__))


def _bootstrap_modules():
    """
    Charge les modules du projet en gérant les imports relatifs
    de calculation_engine (.excel_writer, .number_to_letter_converter).
    """
    # Charger les modules sans imports relatifs
    for mod_name in ['excel_writer', 'number_to_letter_converter', 'data_reader']:
        if mod_name not in sys.modules:
            path = os.path.join(_dir, f'{mod_name}.py')
            spec = importlib.util.spec_from_file_location(mod_name, path)
            mod = importlib.util.module_from_spec(spec)
            sys.modules[mod_name] = mod
            spec.loader.exec_module(mod)

    # Créer un package fictif pour résoudre les imports relatifs de calculation_engine
    if '_estim_pkg' not in sys.modules:
        pkg = types.ModuleType('_estim_pkg')
        pkg.__path__ = [_dir]
        pkg.__package__ = '_estim_pkg'
        sys.modules['_estim_pkg'] = pkg
        sys.modules['_estim_pkg.excel_writer'] = sys.modules['excel_writer']
        sys.modules['_estim_pkg.number_to_letter_converter'] = sys.modules['number_to_letter_converter']

        calc_path = os.path.join(_dir, 'calculation_engine.py')
        spec = importlib.util.spec_from_file_location(
            '_estim_pkg.calculation_engine', calc_path
        )
        calc_mod = importlib.util.module_from_spec(spec)
        calc_mod.__package__ = '_estim_pkg'
        sys.modules['_estim_pkg.calculation_engine'] = calc_mod
        spec.loader.exec_module(calc_mod)


_bootstrap_modules()

from _estim_pkg.calculation_engine import (
    parse_calcul_sheet_and_process_blocks,
    process_menuiserie_block,
    process_simple_block,
    process_formula_block,
    write_recap_block,
)
from data_reader import (
    get_qt_data,
    get_open_data,
    get_simple_block_data,
    get_formula_block_data,
)


def process_estim_batiment(file_bytes):
    """
    Traite le fichier Excel d'estimation et retourne le résultat en mémoire.

    Args:
        file_bytes (bytes): Contenu brut du fichier Excel d'entrée.

    Returns:
        tuple: (output_io, output_filename) en cas de succès,
               (None, message_erreur) en cas d'erreur.
    """
    import openpyxl

    try:
        input_io = io.BytesIO(file_bytes)
        input_wb_formulas = openpyxl.load_workbook(input_io, data_only=False)
        input_io.seek(0)
        input_wb_values = openpyxl.load_workbook(input_io, data_only=True)

        def get_sheet(wb, name):
            return wb[name] if name in wb.sheetnames else None

        qt_sheet        = get_sheet(input_wb_values, "qt")
        calcul_sheet    = get_sheet(input_wb_formulas, "calcul")
        open_sheet      = get_sheet(input_wb_values, "open")
        electricite_sheet = get_sheet(input_wb_values, "Electricite")
        plomberie_sheet = get_sheet(input_wb_values, "Plomberie")
        peinture_sheet  = get_sheet(input_wb_formulas, "Peinture")
        revetement_sheet = get_sheet(input_wb_formulas, "Revetement")
        toiture_sheet   = get_sheet(input_wb_formulas, "Toiture")

        qt_data_dict        = get_qt_data(qt_sheet) if qt_sheet else {}
        open_data_list      = get_open_data(open_sheet) if open_sheet else []
        electricite_data    = get_simple_block_data(electricite_sheet) if electricite_sheet else []
        plomberie_data      = get_simple_block_data(plomberie_sheet) if plomberie_sheet else []
        peinture_data       = get_formula_block_data(peinture_sheet) if peinture_sheet else []
        revetement_data     = get_formula_block_data(revetement_sheet) if revetement_sheet else []
        toiture_data        = get_formula_block_data(toiture_sheet) if toiture_sheet else []

        output_wb = openpyxl.Workbook()
        if "Sheet" in output_wb.sheetnames:
            ws = output_wb["Sheet"]
            ws.title = "Estimation Globale"
        else:
            ws = output_wb.create_sheet("Estimation Globale", 0)

        recap_entries = []
        row = 1

        if calcul_sheet:
            row = parse_calcul_sheet_and_process_blocks(calcul_sheet, qt_data_dict, ws, recap_entries)

        if open_data_list:
            row = process_menuiserie_block(open_data_list, ws, row, recap_entries)

        if electricite_data:
            row = process_simple_block(electricite_data, ws, row, "V", "ELECTRICITE", 1, recap_entries)

        if plomberie_data:
            row = process_simple_block(plomberie_data, ws, row, "VI", "PLOMBERIE SANITAIRE", 1, recap_entries)

        if revetement_data:
            row = process_formula_block(revetement_data, qt_data_dict, ws, row, "VII", "REVETEMENT", 1, recap_entries)

        if peinture_data:
            row = process_formula_block(peinture_data, qt_data_dict, ws, row, "VIII", "PEINTURE", 1, recap_entries)

        if toiture_data:
            row = process_formula_block(toiture_data, qt_data_dict, ws, row, "IX", "TOITURE", 1, recap_entries)

        if recap_entries:
            write_recap_block(ws, row, recap_entries)

        output_io = io.BytesIO()
        output_wb.save(output_io)
        output_io.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"Estimation_Resultat_{timestamp}.xlsx"

        blocs_count = len(recap_entries)
        return output_io, output_filename, blocs_count

    except Exception as exc:
        import traceback
        return None, f"Erreur : {exc}\n{traceback.format_exc()}", 0

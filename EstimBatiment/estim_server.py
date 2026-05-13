# EstimBatiment/estim_server.py
"""
Serveur FastAPI pour EstimBatiment v2.
Lance sur http://127.0.0.1:{ESTIM_PORT} (défaut : 8765)

Endpoints :
  GET  /health
  POST /process           multipart: file=<xlsx> → 5 sorties JSON
  GET  /cache             liste des résultats en cache
  GET  /cache/{ts}/data   toutes les sorties JSON d'un résultat
  DELETE /cache/{ts}      supprime une entrée
  DELETE /cache           vide tout le cache
"""

import os
import sys
import io
import json
import glob
import shutil
from datetime import datetime
from pathlib import Path

_dir = os.path.dirname(os.path.abspath(__file__))
if _dir not in sys.path:
    sys.path.insert(0, _dir)

from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import StreamingResponse
import uvicorn

from new_estim_engine import compute_all
import math

# ─── Config ──────────────────────────────────────────────────────────────────

CACHE_DIR = os.path.join(_dir, ".cache")
os.makedirs(CACHE_DIR, exist_ok=True)

app = FastAPI(title="EstimBatiment API", version="2.0")

# Configuration CORS pour permettre les requêtes depuis Vercel (ou autre domaine front-end)
from fastapi.middleware.cors import CORSMiddleware

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # En production, remplacer "*" par l'URL de l'application Flutter (ex: "https://mon-app.vercel.app")
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def _sanitize(obj):
    """Remplace NaN/Inf par 0 récursivement pour éviter les erreurs JSON."""
    if isinstance(obj, float):
        return 0.0 if (math.isnan(obj) or math.isinf(obj)) else obj
    if isinstance(obj, dict):
        return {k: _sanitize(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_sanitize(v) for v in obj]
    return obj


# ─── Health ──────────────────────────────────────────────────────────────────

@app.get("/health")
def health():
    return {"status": "ok", "version": "2.0"}


# ─── Process ─────────────────────────────────────────────────────────────────

@app.post("/process")
async def process_file(file: UploadFile = File(...)):
    content = await file.read()

    try:
        result = compute_all(content)
        result = _sanitize(result)

        # Sauvegarde en cache (JSON)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        cache_path = os.path.join(CACHE_DIR, f"Estimation_{ts}.json")
        with open(cache_path, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False)

    except Exception as exc:
        import traceback
        raise HTTPException(status_code=422, detail=f"{exc}\n{traceback.format_exc()}")

    return {"timestamp": ts, "outputs": result["outputs"]}


# ─── Cache ───────────────────────────────────────────────────────────────────

@app.get("/cache")
def list_cache():
    files = sorted(
        glob.glob(os.path.join(CACHE_DIR, "Estimation_*.json")),
        key=os.path.getmtime,
        reverse=True,
    )
    entries = []
    for f in files:
        stat = os.stat(f)
        ts = Path(f).stem.replace("Estimation_", "")
        entries.append({
            "timestamp": ts,
            "name": Path(f).name,
            "size": stat.st_size,
            "modified": datetime.fromtimestamp(stat.st_mtime).isoformat(),
        })
    return entries


@app.get("/cache/{ts}/data")
def get_cache_data(ts: str):
    safe_ts = Path(ts).name
    path = os.path.join(CACHE_DIR, f"Estimation_{safe_ts}.json")
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Résultat introuvable")
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    return {"timestamp": safe_ts, "outputs": data["outputs"]}


# ─── Export XLSX ─────────────────────────────────────────────────────────────

def _generate_xlsx(output: dict) -> bytes:
    import io as _io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = output["label"][:31]

    def _cell(row, col, value=None, bg=None, bold=False, white=False, center=False):
        c = ws.cell(row, col, value)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(bold=bold, color="FFFFFF" if white else "1F2937")
        if center:
            c.alignment = Alignment(horizontal="center", vertical="center")
        return c

    cur = 1
    is_mat = output["id"] == "detail_materiaux"

    if is_mat:
        mat_headers = None
        for r in output["rows"]:
            rt = r.get("type", "")
            if rt == "mat_header":
                mat_headers = ["Élément"] + (r.get("headers") or [])
                for ci, h in enumerate(mat_headers, 1):
                    _cell(cur, ci, h, bg="0D47A1", bold=True, white=True, center=True)
                ws.row_dimensions[cur].height = 22
                cur += 1
            elif rt == "mat_section":
                nc = len(mat_headers) if mat_headers else 1
                _cell(cur, 1, r["description"], bg="1565C0", bold=True, white=True)
                for ci in range(2, nc + 1):
                    _cell(cur, ci, bg="1565C0", white=True)
                if nc > 1:
                    ws.merge_cells(start_row=cur, start_column=1,
                                   end_row=cur, end_column=nc)
                cur += 1
            elif rt in ("mat_item", "mat_subtotal", "mat_total"):
                bg_map = {"mat_total": "1B5E20", "mat_subtotal": "1976D2", "mat_item": None}
                bg = bg_map[rt]
                is_bold = bg is not None
                _cell(cur, 1, r["description"], bg=bg, bold=is_bold, white=is_bold)
                for ci, v in enumerate(r.get("values") or [], 2):
                    _cell(cur, ci, v if v else None, bg=bg, bold=is_bold,
                          white=is_bold, center=True)
                cur += 1

        if mat_headers:
            ws.column_dimensions[get_column_letter(1)].width = 38
            for ci in range(2, len(mat_headers) + 1):
                ws.column_dimensions[get_column_letter(ci)].width = 13
    else:
        # Colonnes standard
        hdrs = ["N°", "Désignation", "Unité", "Quantité", "Prix Unit.", "Montant"]
        for ci, h in enumerate(hdrs, 1):
            _cell(cur, ci, h, bg="0D47A1", bold=True, white=True, center=True)
        ws.row_dimensions[cur].height = 22
        cur += 1

        for r in output["rows"]:
            rt = r.get("type", "")
            if rt == "section_hdr":
                for ci in range(1, 7):
                    _cell(cur, ci, bg="0D47A1", bold=True, white=True)
                ws.cell(cur, 1).value = r.get("num", "")
                ws.cell(cur, 2).value = r.get("description", "")
                ws.row_dimensions[cur].height = 20
            elif rt == "item":
                bg = "F0F4F8" if cur % 2 == 0 else None
                for ci in range(1, 7):
                    _cell(cur, ci, bg=bg)
                ws.cell(cur, 1).value = r.get("num", "")
                ws.cell(cur, 2).value = r.get("description", "")
                ws.cell(cur, 3).value = r.get("unite", "")
                ws.cell(cur, 4).value = r.get("quantite")
                ws.cell(cur, 5).value = r.get("pu")
                ws.cell(cur, 6).value = r.get("montant")
            elif rt == "section_total":
                for ci in range(1, 7):
                    c = ws.cell(cur, ci)
                    c.fill = PatternFill("solid", fgColor="BBDEFB")
                    c.font = Font(bold=True, color="0D47A1")
                ws.cell(cur, 1).value = r.get("num", "")
                ws.cell(cur, 6).value = r.get("montant")
            elif rt == "grand_total":
                for ci in range(1, 7):
                    _cell(cur, ci, bg="1B5E20", bold=True, white=True)
                ws.cell(cur, 2).value = r.get("description", "TOTAL GÉNÉRAL HTVA")
                ws.cell(cur, 6).value = r.get("montant")
                ws.row_dimensions[cur].height = 22
            cur += 1

        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 52
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 22

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@app.get("/cache/{ts}/export/{output_id}")
def export_output(ts: str, output_id: str):
    from fastapi.responses import Response as FResponse
    safe_ts = Path(ts).name
    path = os.path.join(CACHE_DIR, f"Estimation_{safe_ts}.json")
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Cache introuvable")
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    output = next((o for o in data.get("outputs", []) if o["id"] == output_id), None)
    if not output:
        raise HTTPException(status_code=404, detail=f"Output '{output_id}' introuvable")
    try:
        xlsx_bytes = _generate_xlsx(output)
    except Exception as exc:
        import traceback
        raise HTTPException(status_code=500, detail=f"{exc}\n{traceback.format_exc()}")
    label = output.get("label", output_id)
    filename = f"{label}_{safe_ts}.xlsx"
    return FResponse(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.delete("/cache/{ts}")
def delete_entry(ts: str):
    safe_ts = Path(ts).name
    path = os.path.join(CACHE_DIR, f"Estimation_{safe_ts}.json")
    if os.path.exists(path):
        os.remove(path)
    return {"deleted": safe_ts}


@app.delete("/cache")
def clear_cache():
    shutil.rmtree(CACHE_DIR, ignore_errors=True)
    os.makedirs(CACHE_DIR, exist_ok=True)
    return {"cleared": True}


# ─── Entry point ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    port = int(os.environ.get("ESTIM_PORT", "8765"))
    print(f"[EstimBatiment] Serveur v2 démarré sur http://127.0.0.1:{port}", flush=True)
    uvicorn.run(app, host="127.0.0.1", port=port, log_level="warning")

# EstimBatiment/estim_server.py
"""
Serveur FastAPI pour EstimBatiment.
Lance sur http://127.0.0.1:{ESTIM_PORT} (défaut : 8765)

Endpoints :
  GET  /health
  POST /process                       multipart: file=<xlsx>
  GET  /cache                         liste des résultats en cache
  GET  /cache/{name}/rows             lignes JSON d'un résultat
  GET  /cache/{name}/download         téléchargement xlsx
  DELETE /cache/{name}                supprime une entrée
  DELETE /cache                       vide tout le cache
"""

import os
import sys
import io
import re
import math
import glob
import shutil
from datetime import datetime
from pathlib import Path

_dir = os.path.dirname(os.path.abspath(__file__))
if _dir not in sys.path:
    sys.path.insert(0, _dir)

from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import StreamingResponse
import uvicorn
import openpyxl

from estim_engine import process_estim_batiment
from detail_engine import compute_detail

# ─── Config ──────────────────────────────────────────────────────────────────

CACHE_DIR = os.path.join(_dir, ".cache")
os.makedirs(CACHE_DIR, exist_ok=True)

ROMAN_SET = {"I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X", "XI", "XII"}

app = FastAPI(title="EstimBatiment API", version="1.0")


# ─── Health ──────────────────────────────────────────────────────────────────

@app.get("/health")
def health():
    return {"status": "ok"}


# ─── Process ─────────────────────────────────────────────────────────────────

@app.post("/process")
async def process_file(file: UploadFile = File(...)):
    content = await file.read()
    output_io, filename_or_error, blocs_count = process_estim_batiment(content)

    if output_io is None:
        raise HTTPException(status_code=422, detail=str(filename_or_error))

    # Sauvegarde en cache
    cache_path = os.path.join(CACHE_DIR, filename_or_error)
    with open(cache_path, "wb") as f:
        f.write(output_io.getvalue())

    # Parse pour affichage
    output_io.seek(0)
    rows = _parse_result_xlsx(output_io)

    return {
        "filename": filename_or_error,
        "blocs_count": blocs_count,
        "rows": rows,
    }


# ─── Process Detail ──────────────────────────────────────────────────────────

@app.post("/process-detail")
async def process_detail(file: UploadFile = File(...)):
    """Calcule le détail des matériaux à partir du fichier EstimType.xlsx."""
    content = await file.read()
    try:
        output_io, filename = compute_detail(content)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=str(exc))

    # Sauvegarde en cache
    cache_path = os.path.join(CACHE_DIR, filename)
    with open(cache_path, "wb") as f:
        f.write(output_io.getvalue())

    # Parse pour affichage
    output_io.seek(0)
    rows = _parse_detail_xlsx(output_io)

    return {
        "filename": filename,
        "rows": rows,
    }


# ─── Cache ───────────────────────────────────────────────────────────────────

@app.get("/cache")
def list_cache():
    files = sorted(
        glob.glob(os.path.join(CACHE_DIR, "*.xlsx")),
        key=os.path.getmtime,
        reverse=True,
    )
    result = []
    for f in files:
        stat = os.stat(f)
        result.append({
            "name": os.path.basename(f),
            "size": stat.st_size,
            "modified": datetime.fromtimestamp(stat.st_mtime).isoformat(),
        })
    return result


@app.get("/cache/{name}/rows")
def get_cache_rows(name: str):
    safe_name = Path(name).name
    path = os.path.join(CACHE_DIR, safe_name)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Fichier introuvable")
    with open(path, "rb") as f:
        rows = _parse_result_xlsx(io.BytesIO(f.read()))
    return {"filename": safe_name, "rows": rows}


@app.get("/cache/{name}/download")
def download_cache(name: str):
    safe_name = Path(name).name
    path = os.path.join(CACHE_DIR, safe_name)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Fichier introuvable")

    def _iter():
        with open(path, "rb") as f:
            while chunk := f.read(65536):
                yield chunk

    return StreamingResponse(
        _iter(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}"'},
    )


@app.delete("/cache/{name}")
def delete_entry(name: str):
    safe_name = Path(name).name
    path = os.path.join(CACHE_DIR, safe_name)
    if os.path.exists(path):
        os.remove(path)
    return {"deleted": safe_name}


@app.delete("/cache")
def clear_cache():
    shutil.rmtree(CACHE_DIR, ignore_errors=True)
    os.makedirs(CACHE_DIR, exist_ok=True)
    return {"cleared": True}


# ─── Parser ──────────────────────────────────────────────────────────────────

def _parse_result_xlsx(xlsx_io) -> list:
    """
    Extrait les lignes du xlsx résultat en dicts JSON.

    Types possibles :
      block_hdr  — entête de bloc
      data       — ligne de données (montant = qté × P.U. calculé ici, pas via formule Excel)
      total      — TOTAL du bloc (accumulé depuis les lignes data)
      recap_hdr  — entête du récapitulatif
      recap_item — ligne du récapitulatif (montant = total du bloc correspondant)
      grand_total — TOTAL GÉNÉRAL (somme des totaux de blocs)
      note       — texte pleine largeur (ex : arrêté du devis)

    NOTE : openpyxl lit None pour les cellules-formule sans cache → les montants
    sont recalculés directement (qté × P.U.) sans relire la colonne F.
    """
    wb = openpyxl.load_workbook(xlsx_io, data_only=True)
    ws = wb.active

    # Résoudre les cellules fusionnées → valeur de la cellule de tête
    merged_values: dict = {}
    for mr in ws.merged_cells.ranges:
        top_val = ws.cell(mr.min_row, mr.min_col).value
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merged_values[(r, c)] = top_val

    def cv(row, col):
        return merged_values.get((row, col), ws.cell(row, col).value)

    rows = []
    current_block_num: str | None = None
    current_block_total: float = 0.0
    block_totals: dict[str, float] = {}   # roman → total calculé
    grand_total_computed: float = 0.0
    seen_notes: set = set()
    in_recap: bool = False                # True après l'entête RÉCAPITULATIF

    for r in range(1, ws.max_row + 1):
        a = cv(r, 1); b = cv(r, 2); c = cv(r, 3)
        d = cv(r, 4); e = cv(r, 5)

        if all(v is None for v in [a, b, c, d, e]):
            continue

        a_str = str(a).strip() if a is not None else ""
        b_str = str(b).strip() if b is not None else ""

        # ── Entête récapitulatif ───────────────────────────────────────────────
        if "RECAPITULATIF" in b_str.upper() or "RÉCAPITULATIF" in b_str.upper():
            in_recap = True
            rows.append({"type": "recap_hdr", "num": "", "description": b_str,
                         "unite": None, "quantite": None, "pu": None, "montant": None})

        # ── TOTAL GÉNÉRAL ──────────────────────────────────────────────────────
        elif "TOTAL GENERAL" in a_str.upper() or "TOTAL GÉNÉRAL" in a_str.upper():
            rows.append({"type": "grand_total", "num": "", "description": a_str,
                         "unite": None, "quantite": None, "pu": None,
                         "montant": grand_total_computed})

        # ── Dans le récapitulatif : chaque bloc = recap_item ──────────────────
        elif in_recap and a_str in ROMAN_SET and b_str:
            recap_total = block_totals.get(a_str, 0.0)
            rows.append({"type": "recap_item", "num": a_str, "description": b_str,
                         "unite": None, "quantite": None, "pu": None,
                         "montant": recap_total})

        # ── Entête de bloc (chiffre romain + titre, hors récap) ───────────────
        elif not in_recap and a_str in ROMAN_SET and b_str and not b_str.startswith("TOTAL"):
            if current_block_num:
                block_totals[current_block_num] = current_block_total
            current_block_num = a_str
            current_block_total = 0.0
            rows.append({"type": "block_hdr", "num": a_str, "description": b_str,
                         "unite": None, "quantite": None, "pu": None, "montant": None})

        # ── Ligne TOTAL du bloc ────────────────────────────────────────────────
        elif a_str.startswith("TOTAL") and "GENERAL" not in a_str.upper() and "GÉNÉRAL" not in a_str:
            if current_block_num:
                block_totals[current_block_num] = current_block_total
                grand_total_computed += current_block_total
            rows.append({"type": "total", "num": a_str, "description": "",
                         "unite": None, "quantite": None, "pu": None,
                         "montant": block_totals.get(current_block_num or "", 0.0)})
            current_block_num = None
            current_block_total = 0.0

        # ── Note pleine largeur (cellule fusionnée A→F, texte long) ───────────
        elif len(a_str) > 30 and (a_str == b_str or b_str == ""):
            if a_str not in seen_notes:
                seen_notes.add(a_str)
                rows.append({"type": "note", "num": "", "description": a_str,
                             "unite": None, "quantite": None, "pu": None, "montant": None})

        # ── Ligne de données ───────────────────────────────────────────────────
        elif b_str and not in_recap:
            qty = _float(d)
            pu_val = _float(e)
            montant = (qty * pu_val) if (qty is not None and pu_val is not None) else None
            if montant is not None:
                current_block_total += montant
            rows.append({"type": "data", "num": a_str, "description": b_str,
                         "unite": str(c).strip() if c is not None else "",
                         "quantite": qty, "pu": pu_val, "montant": montant})

    return rows


def _float(val) -> float | None:
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


# ─── Parser Detail ───────────────────────────────────────────────────────────

def _parse_detail_xlsx(xlsx_io) -> list:
    """Parse un fichier Detail_Materiaux .xlsx en liste de dicts JSON."""
    wb = openpyxl.load_workbook(xlsx_io, data_only=True)
    ws = wb.active

    MAT_LABELS = ['ciment', 'brique', 'hourdi', 'sable', 'granite',
                  'planche', 'terre', 'ha6', 'ha8', 'ha10', 'ha12', 'ha14']
    N_COLS = 16  # A..P

    rows = []
    in_resume = False

    for row in ws.iter_rows(values_only=True):
        if all(c is None for c in row):
            continue
        cells = list(row) + [None] * max(0, N_COLS - len(row))
        a = str(cells[0]).strip() if cells[0] is not None else ""

        # Ignorer titre et entête
        if a in ("DÉTAIL DES MATÉRIAUX", "N°"):
            continue

        # Section header
        if a in {'I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII', 'IX', 'X', 'XI', 'XII'}:
            desc_val = str(cells[1]) if cells[1] else ""
            rows.append({
                "type": "section",
                "num": a,
                "description": desc_val,
                **{k: None for k in MAT_LABELS},
            })
            in_resume = False
            continue

        # TOTAL GÉNÉRAL
        if "TOTAL GÉNÉRAL" in a:
            mat_vals = {}
            for mi, mk in enumerate(MAT_LABELS):
                v = _float(cells[4 + mi])
                mat_vals[mk] = v
            rows.append({
                "type": "total_general",
                "num": "",
                "description": a,
                **mat_vals,
            })
            continue

        # RÉSUMÉ header
        if "RÉSUMÉ" in a.upper():
            in_resume = True
            rows.append({
                "type": "resume_hdr",
                "num": "",
                "description": a,
                **{k: None for k in MAT_LABELS},
            })
            continue

        # RÉSUMÉ items
        if in_resume:
            # Colonnes fusionnées: A..E = label, F..I = total, J..P = result
            total_str = str(cells[5]).strip() if cells[5] is not None else ""
            result_str = str(cells[9]).strip() if cells[9] is not None else ""
            rows.append({
                "type": "resume_item",
                "num": "",
                "description": a,
                "total_display": total_str,
                "result_display": result_str,
                **{k: None for k in MAT_LABELS},
            })
            continue

        # Regular data row
        desc = str(cells[1]) if cells[1] else ""
        unite = str(cells[2]) if cells[2] else ""
        qte = _float(cells[3])
        mat_vals = {}
        for mi, mk in enumerate(MAT_LABELS):
            v = _float(cells[4 + mi])
            mat_vals[mk] = v
        rows.append({
            "type": "data",
            "num": a,
            "description": desc,
            "unite": unite,
            "quantite": qte,
            **mat_vals,
        })

    return rows


@app.get("/cache/{name}/detail-rows")
def get_cache_detail_rows(name: str):
    """Parse un Detail_Materiaux en cache et retourne ses lignes."""
    safe_name = Path(name).name
    path = os.path.join(CACHE_DIR, safe_name)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Fichier introuvable")
    with open(path, "rb") as f:
        rows = _parse_detail_xlsx(io.BytesIO(f.read()))
    return {"filename": safe_name, "rows": rows}


@app.get("/cache/{name}/export-pdf")
def export_pdf(name: str):
    """Convertit un xlsx en cache en PDF simple (tableaux)."""
    safe_name = Path(name).name
    path = os.path.join(CACHE_DIR, safe_name)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Fichier introuvable")

    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
    except ImportError:
        raise HTTPException(status_code=500, detail="reportlab non installé")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Extraire toutes les lignes en texte
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append([str(c) if c is not None else "" for c in row])

    if not data:
        raise HTTPException(status_code=422, detail="Fichier vide")

    # Construire le PDF
    buf = io.BytesIO()
    is_detail = "Detail" in safe_name or "detail" in safe_name
    page_size = landscape(A4) if is_detail else A4
    doc = SimpleDocTemplate(buf, pagesize=page_size,
                            leftMargin=10*mm, rightMargin=10*mm,
                            topMargin=15*mm, bottomMargin=15*mm)

    styles = getSampleStyleSheet()
    elements = []
    title = safe_name.replace(".xlsx", "").replace("_", " ")
    elements.append(Paragraph(f"<b>{title}</b>", styles["Title"]))
    elements.append(Spacer(1, 6*mm))

    # Déterminer largeur colonnes
    n_cols = max(len(r) for r in data) if data else 1
    avail_w = page_size[0] - 20*mm
    col_w = avail_w / n_cols

    table = Table(data, colWidths=[col_w] * n_cols, repeatRows=1)
    style_cmds = [
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.05, 0.28, 0.63)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.97, 1.0)]),
    ]
    table.setStyle(TableStyle(style_cmds))
    elements.append(table)

    doc.build(elements)
    pdf_bytes = buf.getvalue()

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{safe_name.replace(".xlsx", ".pdf")}"'},
    )


# ─── Entry point ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    port = int(os.environ.get("ESTIM_PORT", "8765"))
    print(f"[EstimBatiment] Serveur démarré sur http://127.0.0.1:{port}", flush=True)
    uvicorn.run(app, host="127.0.0.1", port=port, log_level="warning")

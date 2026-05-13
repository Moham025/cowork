# app_ctk.py
import customtkinter as ctk
from tkinter import filedialog, messagebox, ttk
import tkinter as tk
import os, sys, threading, shutil, glob, re
from datetime import datetime

_dir = os.path.dirname(os.path.abspath(__file__))
if _dir not in sys.path:
    sys.path.insert(0, _dir)

CACHE_DIR = os.path.join(_dir, ".cache")
os.makedirs(CACHE_DIR, exist_ok=True)

DEFAULT_FILE  = r"D:\BOLO\9-EstimBatiment\EstimType.xlsx"
ROMAN_SET     = {"I","II","III","IV","V","VI","VII","VIII","IX","X","XI","XII"}

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

WIN_W_CLOSED = 700
WIN_W_OPEN   = 1340
WIN_H        = 640
PANEL_W_INIT = 620
LEFT_MIN     = 420
RIGHT_MIN    = 300

C_HEADER = "#0D47A1"
C_DARK   = "#263238"
C_GREEN  = "#2E7D32"
C_GREEN2 = "#1B5E20"
C_RED    = "#C62828"


# ═══════════════════════════════════════════════════════════════════════════════
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("EstimBatiment")
        self.geometry(f"{WIN_W_CLOSED}x{WIN_H}")
        self.minsize(500, 440)
        self.resizable(True, True)

        self._input_path   = DEFAULT_FILE if os.path.exists(DEFAULT_FILE) else ""
        self._preview_path = None
        self._panel_open   = False

        self._build_ui()
        self._style_treeview()
        self._refresh_cache_list()

    # ══════════════════════════════════════════════════════════════════════════
    def _build_ui(self):
        self.grid_rowconfigure(0, weight=0)
        self.grid_rowconfigure(1, weight=1)
        self.grid_columnconfigure(0, weight=1)

        # ── Header ────────────────────────────────────────────────────────────
        hdr = ctk.CTkFrame(self, fg_color=C_HEADER, corner_radius=0, height=68)
        hdr.grid(row=0, column=0, sticky="ew")
        hdr.grid_propagate(False)
        hdr.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(hdr, text="🏗  EstimBatiment",
                     font=ctk.CTkFont(size=21, weight="bold"),
                     text_color="white").place(x=28, rely=0.35, anchor="w")
        ctk.CTkLabel(hdr, text="Calcul automatique de devis",
                     font=ctk.CTkFont(size=11),
                     text_color="#BBDEFB").place(x=28, rely=0.72, anchor="w")

        # Bouton toggle dans le header (icône flèche uniquement, top-right)
        self._toggle_btn = ctk.CTkButton(
            hdr, text="▶",
            width=38, height=38,
            font=ctk.CTkFont(size=18, weight="bold"),
            fg_color="#1565C0", hover_color="#0D47A1",
            corner_radius=8,
            state="disabled",
            command=self._toggle_panel)
        self._toggle_btn.place(relx=1.0, rely=0.5, anchor="e", x=-16)

        # ── PanedWindow ───────────────────────────────────────────────────────
        self._pw = tk.PanedWindow(
            self, orient="horizontal",
            sashwidth=6, sashrelief="flat",
            bg="#90A4AE", sashpad=0)
        self._pw.grid(row=1, column=0, sticky="nsew")

        # ── Panneau gauche ────────────────────────────────────────────────────
        self._left = ctk.CTkFrame(self._pw, fg_color="#F5F5F5", corner_radius=0)
        self._pw.add(self._left, minsize=LEFT_MIN, stretch="always")
        self._left.grid_columnconfigure(0, weight=1)
        self._left.grid_rowconfigure(5, weight=1)   # cache scroll s'étire

        # -- Fichier source
        ctk.CTkLabel(self._left, text="FICHIER EXCEL SOURCE",
                     font=ctk.CTkFont(size=10, weight="bold"),
                     text_color="#9E9E9E"
                     ).grid(row=0, column=0, sticky="w", padx=28, pady=(20, 4))

        fc = ctk.CTkFrame(self._left, fg_color="white", corner_radius=10)
        fc.grid(row=1, column=0, sticky="ew", padx=28)
        fc.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(fc, text="📊", font=ctk.CTkFont(size=17)
                     ).grid(row=0, column=0, padx=(12, 6), pady=10)
        self._file_label = ctk.CTkLabel(
            fc, text=self._input_path or "Aucun fichier sélectionné",
            font=ctk.CTkFont(size=11), text_color="#424242", anchor="w")
        self._file_label.grid(row=0, column=1, sticky="ew", padx=(0, 6))
        ctk.CTkButton(fc, text="📁  Parcourir", width=118, height=30,
                      fg_color=C_HEADER, hover_color="#0D47A1",
                      font=ctk.CTkFont(size=11),
                      command=self._pick_file
                      ).grid(row=0, column=2, padx=(0, 10), pady=8)

        # -- Calcul
        act = ctk.CTkFrame(self._left, fg_color="transparent")
        act.grid(row=2, column=0, sticky="w", padx=28, pady=(16, 0))
        self._calc_btn = ctk.CTkButton(
            act, text="⚙  Lancer le calcul",
            width=196, height=42,
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=C_HEADER, hover_color="#0D47A1",
            command=self._on_calc_click)
        self._calc_btn.pack(side="left")
        self._spinner = ctk.CTkProgressBar(act, mode="indeterminate", width=140)
        self._spinner.pack(side="left", padx=(12, 0))
        self._spinner.pack_forget()

        # -- Statut
        self._status_var = ctk.StringVar(value="")
        self._status_lbl = ctk.CTkLabel(
            self._left, textvariable=self._status_var,
            font=ctk.CTkFont(size=11), text_color=C_HEADER,
            anchor="w", wraplength=580)
        self._status_lbl.grid(row=3, column=0, sticky="w", padx=28, pady=(6, 0))

        # -- Label cache + bouton vider tout
        cache_hdr = ctk.CTkFrame(self._left, fg_color="transparent")
        cache_hdr.grid(row=4, column=0, sticky="ew", padx=28, pady=(16, 4))
        cache_hdr.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(cache_hdr, text="FICHIERS CALCULÉS (CACHE)",
                     font=ctk.CTkFont(size=10, weight="bold"),
                     text_color="#9E9E9E"
                     ).grid(row=0, column=0, sticky="w")
        ctk.CTkButton(cache_hdr, text="🗑  Vider le cache", width=130, height=24,
                      font=ctk.CTkFont(size=10),
                      fg_color=C_RED, hover_color="#B71C1C",
                      command=self._clear_cache
                      ).grid(row=0, column=1, sticky="e")

        # -- Liste scrollable cache
        self._cache_scroll = ctk.CTkScrollableFrame(
            self._left, fg_color="white", corner_radius=8)
        self._cache_scroll.grid(row=5, column=0, sticky="nsew",
                                padx=28, pady=(0, 18))
        self._cache_scroll.grid_columnconfigure(0, weight=1)

        # ── Panneau droit (aperçu) ────────────────────────────────────────────
        self._right = ctk.CTkFrame(self._pw, fg_color="#ECEFF1", corner_radius=0)
        self._right.grid_rowconfigure(1, weight=1)
        self._right.grid_columnconfigure(0, weight=1)

        rh = ctk.CTkFrame(self._right, fg_color=C_DARK,
                          corner_radius=0, height=42)
        rh.grid(row=0, column=0, sticky="ew")
        rh.grid_propagate(False)
        rh.grid_columnconfigure(0, weight=1)

        self._panel_title = ctk.CTkLabel(
            rh, text="📋  Aperçu du résultat",
            font=ctk.CTkFont(size=12, weight="bold"), text_color="white")
        self._panel_title.place(x=14, rely=0.5, anchor="w")

        # Bouton Détail matériaux
        self._detail_btn = ctk.CTkButton(
            rh, text="📦 Détail", width=90, height=28,
            font=ctk.CTkFont(size=11, weight="bold"),
            fg_color="#37474F", hover_color=C_GREEN,
            text_color="white", state="disabled",
            command=self._open_detail,
        )
        self._detail_btn.place(relx=1.0, rely=0.5, anchor="e", x=-48)

        ctk.CTkButton(
            rh, text="✕", width=30, height=26,
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color="transparent", hover_color="#455A64",
            text_color="white", command=self._toggle_panel
        ).place(relx=1.0, rely=0.5, anchor="e", x=-8)

        # Treeview
        tf = ctk.CTkFrame(self._right, fg_color="white", corner_radius=0)
        tf.grid(row=1, column=0, sticky="nsew")
        tf.grid_rowconfigure(0, weight=1)
        tf.grid_columnconfigure(0, weight=1)

        cols = ("num", "description", "unite", "qte", "pu", "montant")
        self._tree = ttk.Treeview(tf, columns=cols,
                                  show="headings", selectmode="browse")
        for col, lbl, w, stretch in [
            ("num",         "N°",          50,  False),
            ("description", "Description", 240, True),
            ("unite",       "Unité",        54,  False),
            ("qte",         "Quantité",     80,  False),
            ("pu",          "P.U.",         85,  False),
            ("montant",     "Montant",     105,  False),
        ]:
            self._tree.heading(col, text=lbl)
            self._tree.column(col, width=w, minwidth=40, stretch=stretch)

        vsb = ttk.Scrollbar(tf, orient="vertical",   command=self._tree.yview)
        hsb = ttk.Scrollbar(tf, orient="horizontal", command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self._tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        # ── Barre d'export ────────────────────────────────────────────────────
        export_bar = ctk.CTkFrame(self._right, fg_color="#37474F",
                                  corner_radius=0, height=44)
        export_bar.grid(row=2, column=0, sticky="ew")
        export_bar.grid_propagate(False)
        export_bar.grid_columnconfigure(0, weight=1)

        exp_btns = ctk.CTkFrame(export_bar, fg_color="transparent")
        exp_btns.place(relx=0.0, rely=0.5, anchor="w", x=10)

        # CSV
        ctk.CTkButton(
            exp_btns, text="CSV", width=52, height=30,
            font=ctk.CTkFont(size=11, weight="bold"),
            fg_color="#1565C0", hover_color="#0D47A1",
            command=self._export_csv,
        ).pack(side="left", padx=(0, 6))

        # PDF
        ctk.CTkButton(
            exp_btns, text="PDF", width=52, height=30,
            font=ctk.CTkFont(size=11, weight="bold"),
            fg_color="#B71C1C", hover_color="#7F0000",
            command=self._export_pdf,
        ).pack(side="left", padx=(0, 6))

        # TXT
        ctk.CTkButton(
            exp_btns, text="TXT", width=52, height=30,
            font=ctk.CTkFont(size=11, weight="bold"),
            fg_color="#4A148C", hover_color="#2E0057",
            command=self._export_txt,
        ).pack(side="left", padx=(0, 6))

        # JSON
        ctk.CTkButton(
            exp_btns, text="JSON", width=58, height=30,
            font=ctk.CTkFont(size=11, weight="bold"),
            fg_color="#E65100", hover_color="#BF360C",
            command=self._export_json,
        ).pack(side="left", padx=(0, 0))

        # Barre bas (fichier + enregistrer)
        bot = ctk.CTkFrame(self._right, fg_color="#ECEFF1",
                           corner_radius=0, height=44)
        bot.grid(row=3, column=0, sticky="ew")
        bot.grid_propagate(False)
        bot.grid_columnconfigure(0, weight=1)
        self._panel_file_lbl = ctk.CTkLabel(
            bot, text="", font=ctk.CTkFont(size=10),
            text_color="#78909C", anchor="w")
        self._panel_file_lbl.place(x=10, rely=0.5, anchor="w")
        ctk.CTkButton(
            bot, text="⬇  Enregistrer sous…",
            height=30, width=180,
            fg_color=C_GREEN, hover_color=C_GREEN2,
            font=ctk.CTkFont(size=11, weight="bold"),
            command=lambda: self._download_file(self._preview_path)
        ).place(relx=1.0, rely=0.5, anchor="e", x=-10)

    # ══════════════════════════════════════════════════════════════════════════
    def _style_treeview(self):
        s = ttk.Style()
        s.theme_use("clam")
        s.configure("Treeview", background="white", foreground="#212121",
                    rowheight=22, fieldbackground="white",
                    font=("Segoe UI", 10))
        s.configure("Treeview.Heading", background="#1565C0", foreground="white",
                    font=("Segoe UI", 10, "bold"), relief="flat")
        s.map("Treeview.Heading", background=[("active", "#0D47A1")])
        s.map("Treeview",
              background=[("selected", "#BBDEFB")],
              foreground=[("selected", "#0D47A1")])
        self._tree.tag_configure("odd",         background="#FAFAFA")
        self._tree.tag_configure("even",        background="white")
        self._tree.tag_configure("block_hdr",   background="#1565C0",
                                  foreground="white",
                                  font=("Segoe UI", 10, "bold"))
        self._tree.tag_configure("total",       background="#E3F2FD",
                                  foreground="#0D47A1",
                                  font=("Segoe UI", 10, "bold"))
        self._tree.tag_configure("recap",       background=C_DARK,
                                  foreground="white",
                                  font=("Segoe UI", 10, "bold"))
        self._tree.tag_configure("recap_row",   background="#37474F",
                                  foreground="white")
        self._tree.tag_configure("grand_total", background="#1B5E20",
                                  foreground="white",
                                  font=("Segoe UI", 10, "bold"))

    # ══════════════════════════════════════════════════════════════════════════
    def _pick_file(self):
        path = filedialog.askopenfilename(
            title="Sélectionnez le fichier Excel d'estimation",
            initialdir=os.path.dirname(self._input_path) if self._input_path else _dir,
            filetypes=[("Fichiers Excel", "*.xlsx *.xls"),
                       ("Tous les fichiers", "*.*")])
        if path:
            self._input_path = path
            self._file_label.configure(text=path)
            self._set_status("", C_HEADER)

    # ══════════════════════════════════════════════════════════════════════════
    def _on_calc_click(self):
        if not self._input_path or not os.path.exists(self._input_path):
            self._set_status("⚠  Fichier introuvable.", C_RED)
            return
        self._set_status("⏳  Calcul en cours…", C_HEADER)
        self._calc_btn.configure(state="disabled")
        self._spinner.pack(side="left", padx=(12, 0))
        self._spinner.start()
        threading.Thread(target=self._run_calculation, daemon=True).start()

    def _run_calculation(self):
        try:
            from estim_engine import process_estim_batiment
            with open(self._input_path, "rb") as f:
                file_bytes = f.read()
            output_io, filename, blocs_count = process_estim_batiment(file_bytes)
            if output_io:
                ts         = datetime.now().strftime("%Y%m%d_%H%M%S")
                cache_path = os.path.join(CACHE_DIR, f"Estimation_{ts}.xlsx")
                output_io.seek(0)
                with open(cache_path, "wb") as f:
                    f.write(output_io.read())
                self.after(0, lambda: self._on_success(cache_path, blocs_count))
            else:
                self.after(0, lambda: self._on_error(str(filename).split("\n")[0]))
        except Exception as exc:
            self.after(0, lambda: self._on_error(str(exc)))

    def _on_success(self, cache_path, blocs_count):
        self._stop_spinner()
        self._set_status(f"✔  {blocs_count} bloc(s) traité(s) avec succès.", C_GREEN)
        self._toggle_btn.configure(state="normal")
        self._detail_btn.configure(state="normal")
        self._refresh_cache_list()
        self._open_preview(cache_path)

    def _on_error(self, msg):
        self._stop_spinner()
        self._set_status(f"✖  {msg}", C_RED)

    # ══════════════════════════════════════════════════════════════════════════
    #  Liste cache
    # ══════════════════════════════════════════════════════════════════════════
    def _refresh_cache_list(self):
        for w in self._cache_scroll.winfo_children():
            w.destroy()

        files = sorted(glob.glob(os.path.join(CACHE_DIR, "*.xlsx")), reverse=True)

        if not files:
            ctk.CTkLabel(self._cache_scroll,
                         text="Aucun fichier calculé pour l'instant.",
                         font=ctk.CTkFont(size=11), text_color="#BDBDBD"
                         ).grid(row=0, column=0, pady=14, padx=10)
            self._toggle_btn.configure(state="disabled")
            return

        self._toggle_btn.configure(state="normal")
        self._cache_scroll.grid_columnconfigure(0, weight=1)

        for i, fpath in enumerate(files):
            fname = os.path.basename(fpath)
            bg    = "#F5F5F5" if i % 2 == 0 else "white"

            row_fr = ctk.CTkFrame(self._cache_scroll, fg_color=bg,
                                  corner_radius=6, height=36)
            row_fr.grid(row=i, column=0, sticky="ew", padx=4, pady=2)
            row_fr.grid_propagate(False)
            row_fr.grid_columnconfigure(1, weight=1)

            ctk.CTkLabel(row_fr, text="📄", font=ctk.CTkFont(size=13)
                         ).grid(row=0, column=0, padx=(8, 4))
            ctk.CTkLabel(row_fr, text=fname,
                         font=ctk.CTkFont(size=10), text_color="#424242",
                         anchor="w"
                         ).grid(row=0, column=1, sticky="ew", padx=(0, 4))

            btn_fr = ctk.CTkFrame(row_fr, fg_color="transparent")
            btn_fr.grid(row=0, column=2, padx=(0, 6))

            is_detail = fname.startswith("Detail_Materiaux_")
            eye_cmd = (lambda p=fpath: self._show_detail_window(p)) if is_detail \
                      else (lambda p=fpath: self._open_preview(p))
            ctk.CTkButton(btn_fr, text="👁", width=28, height=26,
                          font=ctk.CTkFont(size=13),
                          fg_color=C_HEADER, hover_color="#0D47A1",
                          command=eye_cmd
                          ).pack(side="left", padx=2)
            ctk.CTkButton(btn_fr, text="JSON", width=42, height=26,
                          font=ctk.CTkFont(size=9, weight="bold"),
                          fg_color="#E65100", hover_color="#BF360C",
                          command=lambda p=fpath: self._cache_export_json(p)
                          ).pack(side="left", padx=2)
            ctk.CTkButton(btn_fr, text="PDF", width=36, height=26,
                          font=ctk.CTkFont(size=9, weight="bold"),
                          fg_color="#B71C1C", hover_color="#7F0000",
                          command=lambda p=fpath: self._cache_export_pdf(p)
                          ).pack(side="left", padx=2)
            ctk.CTkButton(btn_fr, text="CSV", width=36, height=26,
                          font=ctk.CTkFont(size=9, weight="bold"),
                          fg_color="#1565C0", hover_color="#0D47A1",
                          command=lambda p=fpath: self._cache_export_csv(p)
                          ).pack(side="left", padx=2)
            ctk.CTkButton(btn_fr, text="⬇", width=28, height=26,
                          font=ctk.CTkFont(size=13),
                          fg_color=C_GREEN, hover_color=C_GREEN2,
                          command=lambda p=fpath, n=fname: self._download_file(p, n)
                          ).pack(side="left", padx=2)
            ctk.CTkButton(btn_fr, text="🗑", width=28, height=26,
                          font=ctk.CTkFont(size=13),
                          fg_color=C_RED, hover_color="#B71C1C",
                          command=lambda p=fpath: self._delete_cache(p)
                          ).pack(side="left", padx=2)

    # ══════════════════════════════════════════════════════════════════════════
    #  Aperçu + chargement tableau
    # ══════════════════════════════════════════════════════════════════════════
    def _open_preview(self, fpath):
        if not fpath or not os.path.exists(fpath):
            return
        self._preview_path = fpath
        self._panel_title.configure(text=f"📋  {os.path.basename(fpath)}")
        self._panel_file_lbl.configure(text=f"Cache : {fpath}")
        self._load_table(fpath)
        if not self._panel_open:
            self._toggle_panel()

    def _load_table(self, fpath):
        import openpyxl

        for row in self._tree.get_children():
            self._tree.delete(row)

        wb = openpyxl.load_workbook(fpath, data_only=True)
        ws = wb.active

        def to_f(v):
            if v is None:
                return None
            if isinstance(v, (int, float)):
                return float(v)
            try:
                return float(str(v).replace(",", ".").replace(" ", ""))
            except Exception:
                return None

        def fmt(v):
            if v is None:
                return ""
            if isinstance(v, (int, float)):
                return f"{v:,.2f}".replace(",", " ").replace(".", ",")
            return str(v)

        # ── Passe 1 : calculer les totaux par bloc ────────────────────────────
        block_totals = {}   # roman -> float
        current_roman  = None
        current_sum    = 0.0

        for r in ws.iter_rows(values_only=True):
            if all(c is None for c in r):
                continue
            cells = list(r) + [None] * max(0, 6 - len(r))
            a = cells[0]; d = cells[3]; e = cells[4]
            a_str = str(a).strip().upper() if a else ""
            b_str = str(cells[1]).strip().upper() if cells[1] else ""

            if "RÉCAPITULATIF" in a_str or "RÉCAPITULATIF" in b_str:
                if current_roman:
                    block_totals[current_roman] = current_sum
                break   # les lignes suivantes sont le récap

            if a_str in ROMAN_SET:
                if current_roman:
                    block_totals[current_roman] = current_sum
                current_roman = a_str
                current_sum   = 0.0
            elif a_str.startswith("TOTAL "):
                roman = a_str.replace("TOTAL ", "").strip()
                block_totals[roman] = current_sum
                current_roman = None
                current_sum   = 0.0
            else:
                qv, pv = to_f(d), to_f(e)
                if qv is not None and pv is not None:
                    current_sum += qv * pv

        if current_roman:
            block_totals[current_roman] = current_sum

        grand_total = sum(block_totals.values())

        # ── Passe 2 : remplir le Treeview ────────────────────────────────────
        in_recap = False
        odd      = True

        for r in ws.iter_rows(values_only=True):
            if all(c is None for c in r):
                continue
            cells = list(r) + [None] * max(0, 6 - len(r))
            a, b, c, d, e, _ = (cells[i] if i < len(cells) else None for i in range(6))

            a_str = str(a).strip().upper() if a else ""
            b_str = str(b).strip().upper() if b else ""

            # ── Déterminer tag et montant affiché ─────────────────────────────
            montant = None

            if "RÉCAPITULATIF" in a_str or "RÉCAPITULATIF" in b_str:
                in_recap = True
                tag      = "recap"

            elif in_recap:
                if "TOTAL GÉNÉRAL" in a_str:
                    montant = grand_total
                    tag     = "grand_total"
                elif a_str in ROMAN_SET:
                    montant = block_totals.get(a_str, 0.0)
                    tag     = "recap_row"
                elif "ARRÊTER" in a_str or "ARRÊTER" in b_str:
                    tag = "odd"
                else:
                    tag = "odd"

            elif a_str in ROMAN_SET:
                tag = "block_hdr"

            elif a_str.startswith("TOTAL "):
                roman   = a_str.replace("TOTAL ", "").strip()
                montant = block_totals.get(roman, 0.0)
                tag     = "total"

            else:
                qv, pv = to_f(d), to_f(e)
                if qv is not None and pv is not None and (qv != 0 or pv != 0):
                    montant = qv * pv
                tag  = "odd" if odd else "even"
                odd  = not odd

            mont_str = fmt(montant) if montant is not None and montant != 0 else ""
            row_data = (fmt(a), fmt(b), fmt(c), fmt(d), fmt(e), mont_str)
            self._tree.insert("", "end", values=row_data, tags=(tag,))

    # ══════════════════════════════════════════════════════════════════════════
    #  Helpers export : récupérer les données du Treeview
    # ══════════════════════════════════════════════════════════════════════════
    def _get_table_data(self):
        """Retourne (headers, rows) depuis le Treeview."""
        headers = ["N°", "Description", "Unité", "Quantité", "P.U.", "Montant"]
        rows = []
        for iid in self._tree.get_children():
            rows.append(list(self._tree.item(iid, "values")))
        return headers, rows

    # ── Export CSV ────────────────────────────────────────────────────────────
    def _export_csv(self):
        if not self._tree.get_children():
            messagebox.showinfo("Export CSV", "Aucune donnée à exporter.")
            return
        dest = filedialog.asksaveasfilename(
            title="Exporter en CSV",
            initialfile="Estimation.csv",
            defaultextension=".csv",
            filetypes=[("Fichiers CSV", "*.csv"), ("Tous les fichiers", "*.*")])
        if not dest:
            return
        import csv
        headers, rows = self._get_table_data()
        try:
            with open(dest, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f, delimiter=";")
                w.writerow(headers)
                w.writerows(rows)
            self._set_status(f"✔  CSV exporté : {dest}", C_GREEN)
        except Exception as ex:
            messagebox.showerror("Erreur CSV", str(ex))

    # ── Export PDF ────────────────────────────────────────────────────────────
    def _export_pdf(self):
        if not self._tree.get_children():
            messagebox.showinfo("Export PDF", "Aucune donnée à exporter.")
            return
        dest = filedialog.asksaveasfilename(
            title="Exporter en PDF",
            initialfile="Estimation.pdf",
            defaultextension=".pdf",
            filetypes=[("Fichiers PDF", "*.pdf"), ("Tous les fichiers", "*.*")])
        if not dest:
            return
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.units import cm
            from reportlab.platypus import (SimpleDocTemplate, Table,
                                             TableStyle, Paragraph, Spacer)
            from reportlab.lib.styles import getSampleStyleSheet

            headers, rows = self._get_table_data()
            fname = os.path.basename(self._preview_path or "Estimation")

            doc = SimpleDocTemplate(dest, pagesize=landscape(A4),
                                    leftMargin=1.5*cm, rightMargin=1.5*cm,
                                    topMargin=1.5*cm, bottomMargin=1.5*cm)
            styles = getSampleStyleSheet()

            # Titre
            title = Paragraph(
                f"<b>Estimation — {fname}</b>",
                styles["Title"])

            # Données tableau
            table_data = [headers] + rows

            # Largeurs colonnes (landscape A4 ≈ 25.7 cm utilisables)
            col_w = [1.2*cm, 9.5*cm, 1.5*cm, 2.5*cm, 2.5*cm, 3.0*cm]

            tbl = Table(table_data, colWidths=col_w, repeatRows=1)

            # Style tableau
            tag_map = {}
            for i, iid in enumerate(self._tree.get_children(), start=1):
                tags = self._tree.item(iid, "tags")
                tag_map[i] = tags[0] if tags else "odd"

            style_cmds = [
                # En-tête
                ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#1565C0")),
                ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
                ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE",    (0, 0), (-1, 0), 8),
                ("ALIGN",       (0, 0), (-1, -1), "LEFT"),
                ("ALIGN",       (3, 1), (5, -1), "RIGHT"),
                ("FONTSIZE",    (0, 1), (-1, -1), 7.5),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                 [colors.HexColor("#FAFAFA"), colors.white]),
                ("GRID",        (0, 0), (-1, -1), 0.3, colors.HexColor("#BDBDBD")),
                ("TOPPADDING",  (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]

            color_map = {
                "block_hdr":   ("#1565C0", "#FFFFFF", True),
                "total":       ("#E3F2FD", "#0D47A1", True),
                "recap":       ("#263238", "#FFFFFF", True),
                "recap_row":   ("#37474F", "#FFFFFF", False),
                "grand_total": ("#1B5E20", "#FFFFFF", True),
            }
            for row_idx, tag in tag_map.items():
                if tag in color_map:
                    bg, fg, bold = color_map[tag]
                    style_cmds.append(("BACKGROUND", (0, row_idx), (-1, row_idx),
                                        colors.HexColor(bg)))
                    style_cmds.append(("TEXTCOLOR",  (0, row_idx), (-1, row_idx),
                                        colors.HexColor(fg)))
                    if bold:
                        style_cmds.append(("FONTNAME", (0, row_idx), (-1, row_idx),
                                            "Helvetica-Bold"))

            tbl.setStyle(TableStyle(style_cmds))

            doc.build([title, Spacer(1, 0.4*cm), tbl])
            self._set_status(f"✔  PDF exporté : {dest}", C_GREEN)
        except Exception as ex:
            messagebox.showerror("Erreur PDF", str(ex))

    # ── Export TXT (copie presse-papier) ──────────────────────────────────────
    def _export_txt(self):
        if not self._tree.get_children():
            messagebox.showinfo("Export TXT", "Aucune donnée à exporter.")
            return

        dest = filedialog.asksaveasfilename(
            title="Exporter en TXT ou copier",
            initialfile="Estimation.txt",
            defaultextension=".txt",
            filetypes=[("Fichiers texte", "*.txt"),
                       ("Tous les fichiers", "*.*")])

        headers, rows = self._get_table_data()
        # Largeurs colonnes pour alignement
        widths = [max(len(str(r[i])) for r in ([headers] + rows))
                  for i in range(len(headers))]

        def fmt_row(r):
            return "  ".join(str(v).ljust(widths[i]) for i, v in enumerate(r))

        sep = "─" * (sum(widths) + 2 * len(widths))
        lines = [sep, fmt_row(headers), sep]
        for r in rows:
            lines.append(fmt_row(r))
        lines.append(sep)
        text = "\n".join(lines)

        if dest:
            try:
                with open(dest, "w", encoding="utf-8") as f:
                    f.write(text)
                self._set_status(f"✔  TXT exporté : {dest}", C_GREEN)
            except Exception as ex:
                messagebox.showerror("Erreur TXT", str(ex))
                return

        # Copier aussi dans le presse-papier
        self.clipboard_clear()
        self.clipboard_append(text)
        if not dest:
            self._set_status("✔  Données copiées dans le presse-papier.", C_GREEN)
        else:
            self._set_status(
                f"✔  TXT exporté et copié dans le presse-papier : {dest}", C_GREEN)

    # ── Export JSON ───────────────────────────────────────────────────────────
    def _export_json(self):
        if not self._tree.get_children():
            messagebox.showinfo("Export JSON", "Aucune donnée à exporter.")
            return
        dest = filedialog.asksaveasfilename(
            title="Exporter en JSON",
            initialfile="Estimation.json",
            defaultextension=".json",
            filetypes=[("Fichiers JSON", "*.json"), ("Tous les fichiers", "*.*")])
        if not dest:
            return
        import json
        headers, rows = self._get_table_data()
        keys = ["numero", "description", "unite", "quantite", "prix_unitaire", "montant"]
        data = []
        for row in rows:
            entry = {keys[i]: row[i] for i in range(min(len(keys), len(row)))}
            data.append(entry)
        export = {
            "fichier_source": os.path.basename(self._preview_path or ""),
            "lignes": data,
        }
        try:
            with open(dest, "w", encoding="utf-8") as f:
                json.dump(export, f, ensure_ascii=False, indent=2)
            self._set_status(f"✔  JSON exporté : {dest}", C_GREEN)
        except Exception as ex:
            messagebox.showerror("Erreur JSON", str(ex))

    # ══════════════════════════════════════════════════════════════════════════
    #  Détail des matériaux
    # ══════════════════════════════════════════════════════════════════════════
    def _open_detail(self):
        if not self._input_path or not os.path.exists(self._input_path):
            messagebox.showwarning("Fichier absent",
                                   "Le fichier source est introuvable.")
            return

        # Extraire le timestamp de l'estimation en cours
        ts = None
        if self._preview_path:
            m = re.match(r'Estimation_(\d{8}_\d{6})\.xlsx',
                         os.path.basename(self._preview_path))
            if m:
                ts = m.group(1)

        # Si le détail est déjà en cache, l'afficher directement
        if ts:
            detail_path = os.path.join(CACHE_DIR, f"Detail_Materiaux_{ts}.xlsx")
            if os.path.exists(detail_path):
                self._show_detail_window(detail_path)
                return

        self._detail_btn.configure(state="disabled", text="⏳ Calcul…")
        self.update()

        try:
            from detail_engine import compute_detail
            with open(self._input_path, "rb") as f:
                file_bytes = f.read()
            output_io, _ = compute_detail(file_bytes)

            # Nommer le fichier avec le même timestamp que l'estimation
            if ts:
                fname = f"Detail_Materiaux_{ts}.xlsx"
            else:
                fname = f"Detail_Materiaux_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            cache_path = os.path.join(CACHE_DIR, fname)
            output_io.seek(0)
            with open(cache_path, "wb") as f:
                f.write(output_io.read())

            self._refresh_cache_list()
            self._detail_btn.configure(state="normal", text="📦 Détail")
            self._show_detail_window(cache_path)

        except Exception as ex:
            self._detail_btn.configure(state="normal", text="📦 Détail")
            messagebox.showerror("Erreur Détail", str(ex))

    def _show_detail_window(self, fpath):
        import openpyxl

        win = ctk.CTkToplevel(self)
        win.title("Détail des Matériaux")
        win.geometry("1400x700")
        win.minsize(900, 500)

        # ── Header ────────────────────────────────────────────────────────────
        hdr = ctk.CTkFrame(win, fg_color=C_GREEN, corner_radius=0, height=52)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        hdr.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(hdr, text="📦  Détail des Matériaux — Approvisionnement",
                     font=ctk.CTkFont(size=14, weight="bold"),
                     text_color="white").place(x=16, rely=0.5, anchor="w")

        def save_detail():
            dest = filedialog.asksaveasfilename(
                title="Enregistrer le détail",
                initialfile=os.path.basename(fpath),
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx"), ("Tous", "*.*")])
            if dest:
                shutil.copy2(fpath, dest)

        ctk.CTkButton(hdr, text="⬇  Enregistrer", width=150, height=32,
                      fg_color="#1B5E20", hover_color="#0D3017",
                      font=ctk.CTkFont(size=11, weight="bold"),
                      command=save_detail
                      ).place(relx=1.0, rely=0.5, anchor="e", x=-12)

        # ── Treeview ──────────────────────────────────────────────────────────
        tf = ctk.CTkFrame(win, fg_color="white", corner_radius=0)
        tf.pack(fill="both", expand=True)
        tf.grid_rowconfigure(0, weight=1)
        tf.grid_columnconfigure(0, weight=1)

        cols_def = [
            ("num",      "N°",         55,  False),
            ("desc",     "Description",260, True),
            ("unite",    "Unité",       55,  False),
            ("qte",      "Quantité",    75,  False),
            ("ciment",   "Ciment\n(sacs)", 80, False),
            ("brique",   "Brique\n(nb)",   75, False),
            ("hourdi",   "Hourdi\n(nb)",   75, False),
            ("sable",    "Sable\n(m³)",    75, False),
            ("granite",  "Granite\n(m³)",  75, False),
            ("planche",  "Planche\n(m²)",  75, False),
            ("terre",    "Terre\n(m³)",    75, False),
            ("ha6",      "Ha6\n(ml)",      70, False),
            ("ha8",      "Ha8\n(ml)",      70, False),
            ("ha10",     "Ha10\n(ml)",     70, False),
            ("ha12",     "Ha12\n(ml)",     70, False),
            ("ha14",     "Ha14\n(ml)",     70, False),
        ]
        col_ids = [c[0] for c in cols_def]

        tree = ttk.Treeview(tf, columns=col_ids,
                            show="headings", selectmode="browse")
        for cid, lbl, w, stretch in cols_def:
            tree.heading(cid, text=lbl)
            tree.column(cid, width=w, minwidth=40, stretch=stretch,
                        anchor="e" if cid not in ("num", "desc", "unite") else "center")

        vsb = ttk.Scrollbar(tf, orient="vertical",   command=tree.yview)
        hsb = ttk.Scrollbar(tf, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        # Style treeview dédié
        style = ttk.Style()
        tree.tag_configure("sec",   background="#1565C0", foreground="white")
        tree.tag_configure("total", background="#263238", foreground="white")
        tree.tag_configure("res",   background="#2E7D32", foreground="white")
        tree.tag_configure("r_odd", background="#E8F5E9")
        tree.tag_configure("r_ev",  background="#FFFFFF")
        tree.tag_configure("odd",   background="#FAFAFA")
        tree.tag_configure("even",  background="#FFFFFF")

        # Charger le fichier Excel généré
        wb = openpyxl.load_workbook(fpath, data_only=True)
        ws2 = wb.active

        in_recap = False
        odd = True
        recap_odd = True

        def fv(v):
            if v is None: return ""
            if isinstance(v, (int, float)): return f"{v:,.2f}".replace(",", " ").replace(".", ",")
            return str(v)

        N_COLS_DETAIL = len(cols_def)
        for row in ws2.iter_rows(values_only=True):
            if all(c is None for c in row):
                continue
            cells = list(row) + [None] * max(0, N_COLS_DETAIL - len(row))
            a = str(cells[0]).strip() if cells[0] is not None else ""

            # Ignorer ligne titre ("DÉTAIL DES MATÉRIAUX") et ligne entête ("N°")
            if a in ("DÉTAIL DES MATÉRIAUX", "N°"):
                continue

            # Section header (col A = roman)
            if a in {'I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII', 'IX', 'X'}:
                tree.insert("", "end",
                            values=(a,) + tuple(fv(c) for c in cells[1:N_COLS_DETAIL]),
                            tags=("sec",))
                in_recap = False
                odd = True
                continue

            if "TOTAL GÉNÉRAL" in a or "RÉSUMÉ" in a:
                in_recap = "RÉSUMÉ" in a
                tag = "total" if "TOTAL" in a else "res"
                tree.insert("", "end",
                            values=tuple(fv(c) for c in cells[:N_COLS_DETAIL]),
                            tags=(tag,))
                continue

            if in_recap:
                tag = "r_odd" if recap_odd else "r_ev"
                recap_odd = not recap_odd
            else:
                tag = "odd" if odd else "even"
                odd = not odd

            tree.insert("", "end",
                        values=tuple(fv(c) for c in cells[:N_COLS_DETAIL]),
                        tags=(tag,))

        win.lift()
        win.focus_force()

    # ══════════════════════════════════════════════════════════════════════════
    #  Helpers export depuis le cache
    # ══════════════════════════════════════════════════════════════════════════
    def _read_estim_xlsx(self, fpath):
        """Lit un xlsx d'estimation et retourne (headers, rows) avec montant calculé."""
        import openpyxl

        def to_f(v):
            if v is None: return None
            if isinstance(v, (int, float)): return float(v)
            try: return float(str(v).replace(",", ".").replace(" ", ""))
            except: return None

        def fmt(v):
            if v is None: return ""
            if isinstance(v, (int, float)):
                return f"{v:,.2f}".replace(",", " ").replace(".", ",")
            return str(v)

        wb = openpyxl.load_workbook(fpath, data_only=True)
        ws = wb.active

        # Passe 1 : totaux par bloc
        block_totals = {}
        current_roman = None
        current_sum = 0.0
        for r in ws.iter_rows(values_only=True):
            if all(c is None for c in r): continue
            cells = list(r) + [None] * max(0, 6 - len(r))
            a = cells[0]; d = cells[3]; e = cells[4]
            a_str = str(a).strip().upper() if a else ""
            b_str = str(cells[1]).strip().upper() if cells[1] else ""
            if "RÉCAPITULATIF" in a_str or "RÉCAPITULATIF" in b_str:
                if current_roman: block_totals[current_roman] = current_sum
                break
            if a_str in ROMAN_SET:
                if current_roman: block_totals[current_roman] = current_sum
                current_roman = a_str; current_sum = 0.0
            elif a_str.startswith("TOTAL "):
                roman = a_str.replace("TOTAL ", "").strip()
                block_totals[roman] = current_sum
                current_roman = None; current_sum = 0.0
            else:
                qv, pv = to_f(d), to_f(e)
                if qv is not None and pv is not None:
                    current_sum += qv * pv
        if current_roman:
            block_totals[current_roman] = current_sum
        grand_total = sum(block_totals.values())

        # Passe 2 : collecter les lignes
        headers = ["N°", "Description", "Unité", "Quantité", "Prix Unitaire", "Montant"]
        rows = []
        in_recap = False
        odd = True
        for r in ws.iter_rows(values_only=True):
            if all(c is None for c in r): continue
            cells = list(r) + [None] * max(0, 6 - len(r))
            a, b, c, d, e, _ = (cells[i] if i < len(cells) else None for i in range(6))
            a_str = str(a).strip().upper() if a else ""
            b_str = str(b).strip().upper() if b else ""
            montant = None
            if "RÉCAPITULATIF" in a_str or "RÉCAPITULATIF" in b_str:
                in_recap = True
            elif in_recap:
                if "TOTAL GÉNÉRAL" in a_str:
                    montant = grand_total
                elif a_str in ROMAN_SET:
                    montant = block_totals.get(a_str, 0.0)
            elif a_str in ROMAN_SET:
                pass
            elif a_str.startswith("TOTAL "):
                roman = a_str.replace("TOTAL ", "").strip()
                montant = block_totals.get(roman, 0.0)
            else:
                qv, pv = to_f(d), to_f(e)
                if qv is not None and pv is not None and (qv != 0 or pv != 0):
                    montant = qv * pv
                odd = not odd
            mont_str = fmt(montant) if montant is not None and montant != 0 else ""
            rows.append([fmt(a), fmt(b), fmt(c), fmt(d), fmt(e), mont_str])
        return headers, rows

    def _read_detail_resume_xlsx(self, fpath):
        """Lit un xlsx Detail_Materiaux et retourne les items du résumé approvisionnement."""
        import openpyxl
        wb = openpyxl.load_workbook(fpath, data_only=True)
        ws = wb.active
        resume_items = []
        in_resume = False
        for row in ws.iter_rows(values_only=True):
            if all(c is None for c in row): continue
            a = row[0]
            a_str = str(a).strip() if a is not None else ""
            if "RÉSUMÉ" in a_str.upper() and "APPROVISIONNEMENT" in a_str.upper():
                in_resume = True
                continue
            if in_resume and a_str:
                total_str  = str(row[5]).strip()  if len(row) > 5  and row[5]  is not None else ""
                result_str = str(row[9]).strip()  if len(row) > 9  and row[9]  is not None else ""
                resume_items.append({
                    "materiau": a_str,
                    "total":    total_str,
                    "commande": result_str,
                })
        return resume_items

    # ── Export JSON depuis le cache ───────────────────────────────────────────
    def _cache_export_json(self, fpath):
        if not os.path.exists(fpath):
            messagebox.showwarning("Fichier absent", "Le fichier n'existe plus en cache.")
            return
        fname = os.path.basename(fpath)
        is_detail = fname.startswith("Detail_Materiaux_")
        dest = filedialog.asksaveasfilename(
            title="Exporter en JSON",
            initialfile=fname.replace(".xlsx", ".json"),
            defaultextension=".json",
            filetypes=[("Fichiers JSON", "*.json"), ("Tous les fichiers", "*.*")])
        if not dest: return
        import json
        try:
            if is_detail:
                items = self._read_detail_resume_xlsx(fpath)
                export = {"fichier_source": fname, "resume_approvisionnement": items}
            else:
                _, rows = self._read_estim_xlsx(fpath)
                keys = ["numero", "description", "unite", "quantite", "prix_unitaire", "montant"]
                lignes = [{keys[i]: row[i] for i in range(min(len(keys), len(row)))} for row in rows]
                export = {"fichier_source": fname, "lignes": lignes}
            with open(dest, "w", encoding="utf-8") as f:
                json.dump(export, f, ensure_ascii=False, indent=2)
            self._set_status(f"✔  JSON exporté : {dest}", C_GREEN)
        except Exception as ex:
            messagebox.showerror("Erreur JSON", str(ex))

    # ── Export CSV depuis le cache ────────────────────────────────────────────
    def _cache_export_csv(self, fpath):
        if not os.path.exists(fpath):
            messagebox.showwarning("Fichier absent", "Le fichier n'existe plus en cache.")
            return
        fname = os.path.basename(fpath)
        is_detail = fname.startswith("Detail_Materiaux_")
        dest = filedialog.asksaveasfilename(
            title="Exporter en CSV",
            initialfile=fname.replace(".xlsx", ".csv"),
            defaultextension=".csv",
            filetypes=[("Fichiers CSV", "*.csv"), ("Tous les fichiers", "*.*")])
        if not dest: return
        import csv
        import openpyxl as _xl
        try:
            if is_detail:
                wb = _xl.load_workbook(fpath, data_only=True)
                ws = wb.active
                headers = ["N°", "Description", "Unité", "Quantité",
                           "Ciment(sacs)", "Brique(nb)", "Hourdi(nb)",
                           "Sable(m³)", "Granite(m³)", "Planche(m²)", "Terre(m³)",
                           "Ha6(ml)", "Ha8(ml)", "Ha10(ml)", "Ha12(ml)", "Ha14(ml)"]
                n_cols = len(headers)
                def fv(v):
                    if v is None: return ""
                    if isinstance(v, (int, float)):
                        return f"{v:,.2f}".replace(",", " ").replace(".", ",")
                    return str(v)
                with open(dest, "w", newline="", encoding="utf-8-sig") as f:
                    w = csv.writer(f, delimiter=";")
                    w.writerow(headers)
                    for row in ws.iter_rows(values_only=True):
                        if all(c is None for c in row): continue
                        a = str(row[0]).strip() if row[0] else ""
                        if a in ("DÉTAIL DES MATÉRIAUX", "N°"): continue
                        cells = list(row) + [None] * max(0, n_cols - len(row))
                        w.writerow([fv(cells[i]) for i in range(n_cols)])
            else:
                headers, rows = self._read_estim_xlsx(fpath)
                with open(dest, "w", newline="", encoding="utf-8-sig") as f:
                    w = csv.writer(f, delimiter=";")
                    w.writerow(headers)
                    w.writerows(rows)
            self._set_status(f"✔  CSV exporté : {dest}", C_GREEN)
        except Exception as ex:
            messagebox.showerror("Erreur CSV", str(ex))

    # ── Export PDF depuis le cache ────────────────────────────────────────────
    def _cache_export_pdf(self, fpath):
        if not os.path.exists(fpath):
            messagebox.showwarning("Fichier absent", "Le fichier n'existe plus en cache.")
            return
        fname = os.path.basename(fpath)
        is_detail = fname.startswith("Detail_Materiaux_")
        dest = filedialog.asksaveasfilename(
            title="Exporter en PDF",
            initialfile=fname.replace(".xlsx", ".pdf"),
            defaultextension=".pdf",
            filetypes=[("Fichiers PDF", "*.pdf"), ("Tous les fichiers", "*.*")])
        if not dest: return
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.units import cm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            import openpyxl as _xl

            styles = getSampleStyleSheet()

            if is_detail:
                wb = _xl.load_workbook(fpath, data_only=True)
                ws = wb.active
                headers = ["N°", "Description", "Unité", "Qté",
                           "Ciment\n(sacs)", "Brique\n(nb)", "Hourdi\n(nb)",
                           "Sable\n(m³)", "Granite\n(m³)", "Planche\n(m²)", "Terre\n(m³)",
                           "Ha6\n(ml)", "Ha8\n(ml)", "Ha10\n(ml)", "Ha12\n(ml)", "Ha14\n(ml)"]
                n_cols = 16
                col_w = [0.6*cm, 5.0*cm, 0.7*cm, 1.0*cm] + [1.15*cm] * 12
                def fv(v):
                    if v is None: return ""
                    if isinstance(v, (int, float)):
                        return f"{v:,.2f}".replace(",", " ").replace(".", ",")
                    return str(v)
                rows = []
                for row in ws.iter_rows(values_only=True):
                    if all(c is None for c in row): continue
                    a = str(row[0]).strip() if row[0] else ""
                    if a in ("DÉTAIL DES MATÉRIAUX", "N°"): continue
                    cells = list(row) + [None] * max(0, n_cols - len(row))
                    rows.append([fv(cells[i]) for i in range(n_cols)])
                title_txt = f"<b>Détail des Matériaux — {fname}</b>"
            else:
                headers, rows = self._read_estim_xlsx(fpath)
                col_w = [1.2*cm, 9.5*cm, 1.5*cm, 2.5*cm, 2.5*cm, 3.0*cm]
                title_txt = f"<b>Estimation — {fname}</b>"

            title = Paragraph(title_txt, styles["Title"])
            table_data = [headers] + rows
            tbl = Table(table_data, colWidths=col_w, repeatRows=1)
            style_cmds = [
                ("BACKGROUND",    (0, 0), (-1, 0), colors.HexColor("#1565C0")),
                ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
                ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE",      (0, 0), (-1, 0), 7),
                ("ALIGN",         (0, 0), (-1, -1), "LEFT"),
                ("FONTSIZE",      (0, 1), (-1, -1), 6.5),
                ("ROWBACKGROUNDS",(0, 1), (-1, -1),
                 [colors.HexColor("#FAFAFA"), colors.white]),
                ("GRID",          (0, 0), (-1, -1), 0.3, colors.HexColor("#BDBDBD")),
                ("TOPPADDING",    (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ]
            tbl.setStyle(TableStyle(style_cmds))
            doc = SimpleDocTemplate(dest, pagesize=landscape(A4),
                                    leftMargin=1.0*cm, rightMargin=1.0*cm,
                                    topMargin=1.5*cm, bottomMargin=1.5*cm)
            doc.build([title, Spacer(1, 0.3*cm), tbl])
            self._set_status(f"✔  PDF exporté : {dest}", C_GREEN)
        except Exception as ex:
            messagebox.showerror("Erreur PDF", str(ex))

    # ══════════════════════════════════════════════════════════════════════════
    def _download_file(self, fpath, fname=None):
        if not fpath or not os.path.exists(fpath):
            messagebox.showwarning("Fichier absent",
                                   "Le fichier n'existe plus en cache.")
            return
        dest = filedialog.asksaveasfilename(
            title="Enregistrer le fichier calculé",
            initialfile=fname or os.path.basename(fpath),
            defaultextension=".xlsx",
            filetypes=[("Fichiers Excel", "*.xlsx"),
                       ("Tous les fichiers", "*.*")])
        if not dest:
            return
        try:
            shutil.copy2(fpath, dest)
            self._set_status(f"✔  Enregistré : {dest}", C_GREEN)
        except Exception as ex:
            messagebox.showerror("Erreur", f"Impossible d'enregistrer :\n{ex}")

    def _delete_cache(self, fpath):
        if not messagebox.askyesno("Supprimer",
                                   f"Supprimer « {os.path.basename(fpath)} » du cache ?"):
            return
        try:
            os.remove(fpath)
        except Exception:
            pass
        if self._preview_path == fpath:
            self._preview_path = None
            for row in self._tree.get_children():
                self._tree.delete(row)
            self._panel_title.configure(text="📋  Aperçu du résultat")
            self._panel_file_lbl.configure(text="")
        self._refresh_cache_list()

    def _clear_cache(self):
        files = glob.glob(os.path.join(CACHE_DIR, "*.xlsx"))
        if not files:
            messagebox.showinfo("Cache vide", "Le cache est déjà vide.")
            return
        if not messagebox.askyesno("Vider le cache",
                                   f"Supprimer les {len(files)} fichier(s) du cache ?"):
            return
        for f in files:
            try:
                os.remove(f)
            except Exception:
                pass
        self._preview_path = None
        for row in self._tree.get_children():
            self._tree.delete(row)
        self._panel_title.configure(text="📋  Aperçu du résultat")
        self._panel_file_lbl.configure(text="")
        self._refresh_cache_list()
        self._set_status("✔  Cache vidé.", C_GREEN)

    # ══════════════════════════════════════════════════════════════════════════
    #  Toggle panneau (PanedWindow add/forget)
    # ══════════════════════════════════════════════════════════════════════════
    def _toggle_panel(self):
        if not self._panel_open:
            self._panel_open = True
            self._toggle_btn.configure(text="◀")
            self.geometry(f"{WIN_W_OPEN}x{self.winfo_height()}")
            self._pw.add(self._right, minsize=RIGHT_MIN, stretch="always")
            self.update_idletasks()
            try:
                self._pw.sash_place(0, WIN_W_OPEN - PANEL_W_INIT, 0)
            except Exception:
                pass
        else:
            self._panel_open = False
            self._toggle_btn.configure(text="▶")
            self._pw.forget(self._right)
            self.geometry(f"{WIN_W_CLOSED}x{self.winfo_height()}")

    # ══════════════════════════════════════════════════════════════════════════
    def _stop_spinner(self):
        self._spinner.stop()
        self._spinner.pack_forget()
        self._calc_btn.configure(state="normal")

    def _set_status(self, msg, color):
        self._status_var.set(msg)
        self._status_lbl.configure(text_color=color)


if __name__ == "__main__":
    app = App()
    app.mainloop()

# calculation_engine.py
import re
# Import de la fonction d'écriture Excel avec un import relatif (ajout du '.')
from .excel_writer import create_excel_table_for_block
# Import de la fonction de conversion nombre-lettre avec un import relatif (ajout du '.')
from .number_to_letter_converter import conv_number_letter
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

def evaluate_formula(formula_str, qt_data, current_item_description="N/A"):
    """
    Evaluates a given mathematical formula, replacing 'qt' sheet item references
    with their numerical values. Handles format conversions and errors.
    """
    if not isinstance(formula_str, str) or not formula_str.strip():
        # If the formula is empty or not a string, return the value if it's numeric
        if isinstance(formula_str, (int, float)):
            return float(formula_str)
        return 0.0

    original_formula = formula_str
    
    # Replace commas with periods for Python compatibility
    formula_processed = original_formula.replace(',', '.')
    ops_and_parens = ['+', '-', '*', '/', '(', ')']
    # Add spaces around operators and parentheses for easier tokenization
    for op in ops_and_parens:
        formula_processed = formula_processed.replace(op, f" {op} ")
    formula_processed = ' '.join(formula_processed.split()) # Clean up multiple spaces
    
    raw_tokens = formula_processed.split(' ')
    tokens = [t for t in raw_tokens if t] # Remove empty tokens

    processed_tokens = []
    for token_str in tokens:
        # Try to match an item reference (e.g., "LONGRINE[ml]")
        match_item_header = re.fullmatch(r"([a-zA-Z0-9_ÉÈÀÊÛÔÎÇ.\s-]+)\[([a-zA-Z0-9_]+)\]", token_str, re.IGNORECASE)
        
        if match_item_header:
            item_name_formula = match_item_header.group(1).strip().lower()
            header_name_formula = match_item_header.group(2).strip().lower()
            
            if item_name_formula in qt_data:
                if header_name_formula in qt_data[item_name_formula]:
                    value = qt_data[item_name_formula][header_name_formula]
                    if value is None:
                        # If value is None, use 0.0 and print a warning
                        print(f"    WARNING [Token ITEM] ({current_item_description}): Missing value (None) for '{item_name_formula}[{header_name_formula}]' in 'qt'. Using 0.0.")
                        processed_tokens.append("0.0")
                    elif isinstance(value, (int, float)):
                        # Add the numerical value
                        processed_tokens.append(str(float(value)))
                    else:
                        # If value is of an unexpected type, use 0.0
                        print(f"    WARNING [Token ITEM] ({current_item_description}): Non-numeric value '{value}' for '{item_name_formula}[{header_name_formula}]'. Using 0.0.")
                        processed_tokens.append("0.0")
                else:
                    print(f"    ERROR [Token ITEM] ({current_item_description}): Header '{header_name_formula}' not found for item '{item_name_formula}' in 'qt'. Formula: {original_formula}")
                    return None
            else:
                print(f"    ERROR [Token ITEM] ({current_item_description}): Item '{item_name_formula}' not found in 'qt' data. Formula: {original_formula}")
                return None
        elif re.fullmatch(r"-?\d+(\.\d+)?", token_str):
            # It's a number (integer or decimal)
            processed_tokens.append(token_str)
        elif token_str in ops_and_parens:
            # It's an operator or parenthesis
            processed_tokens.append(token_str)
        else:
            print(f"    ERROR [Token UNKNOWN] ({current_item_description}): Unrecognized token: '{token_str}' in formula '{original_formula}'.")
            return None

    final_expression = " ".join(processed_tokens)

    if not final_expression.strip():
        return 0.0

    try:
        temp_expr_check = final_expression.lower()
        for char in temp_expr_check:
            if not (char.isdigit() or char == '.' or char in ops_and_parens or char.isspace() or char == 'e'):
                print(f"  ERROR [evaluate_formula] ({current_item_description}): Expression '{final_expression}' contains unauthorized character '{char}'. Original: '{original_formula}'")
                return None

        result = eval(final_expression, {"__builtins__": {}}, {})
        return float(result)
    except SyntaxError as e_syn:
        print(f"  ERROR Syntax [evaluate_formula] ({current_item_description}): Incorrect syntax in '{final_expression}'. Original: '{original_formula}'. Error: {e_syn}")
        return None
    except ZeroDivisionError:
        print(f"  ERROR Division by zero [evaluate_formula] ({current_item_description}): In '{final_expression}'. Original: '{original_formula}'")
        return 0.0
    except Exception as e:
        print(f"  ERROR Unexpected [evaluate_formula] ({current_item_description}): While evaluating '{final_expression}' (original: '{original_formula}'): {e}")
        return None


def parse_calcul_sheet_and_process_blocks(calcul_sheet, qt_data, output_ws, recap_entries):
    """
    Parses the 'calcul' sheet, identifies blocks and their items,
    evaluates quantities, and generates tables on the single output sheet.
    
    Args:
        calcul_sheet (openpyxl.worksheet.worksheet.Worksheet): The 'calcul' sheet to read.
        qt_data (dict): The quantity data from the 'qt' sheet.
        output_ws (openpyxl.worksheet.worksheet.Worksheet): The output sheet to write results to.
        recap_entries (list): A list to append recap data (roman_numeral, title, total_cell_ref, numeric_total).
    
    Returns:
        int: The next available row number after processing all blocks from 'calcul' sheet.
    """
    current_block_roman = None
    current_block_title = None
    current_block_items = []
    
    current_excel_row = 1 

    max_row_to_iterate = calcul_sheet.max_row
    print(f"Starting analysis of 'calcul' sheet up to row {max_row_to_iterate}.")

    rows_data = list(calcul_sheet.iter_rows(min_row=1, max_row=max_row_to_iterate, values_only=True))

    for i_row, row_values in enumerate(rows_data):
        col_a_val = str(row_values[0]).strip() if row_values and row_values[0] is not None else None
        col_b_val = row_values[1] if row_values and len(row_values) > 1 else None
        col_c_val = row_values[2] if row_values and len(row_values) > 2 else None
        col_d_val_formula = row_values[3] if row_values and len(row_values) > 3 else None
        col_e_val_pu = row_values[4] if row_values and len(row_values) > 4 else None

        if all(v is None for v in [col_a_val, col_b_val, col_c_val, col_d_val_formula, col_e_val_pu]) and i_row > 0:
             continue

        is_roman = col_a_val is not None and col_a_val.upper() in ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X", "XI", "XII", "XIII", "XIV", "XV"]
        
        if is_roman:
            if current_block_roman and current_block_items:
                print(f"\nProcessing previous block: {current_block_roman} - {current_block_title} with {len(current_block_items)} items.")
                processed_items_for_table = []
                for item in current_block_items:
                    desc, unit, formula, pu_raw = item
                    qty_calculated = evaluate_formula(formula, qt_data, desc)
                    
                    unit_price = 0.0
                    if pu_raw is not None:
                        try:
                            if isinstance(pu_raw, str):
                                unit_price = float(pu_raw.replace(',','.'))
                            elif isinstance(pu_raw, (int, float)):
                                unit_price = float(pu_raw)
                        except ValueError:
                            print(f"WARNING: Invalid unit price '{pu_raw}' for '{desc}'. Using 0.0.")
                            unit_price = 0.0
                    
                    processed_items_for_table.append([desc, unit, qty_calculated if qty_calculated is not None else 0.0, unit_price])

                if processed_items_for_table:
                     next_row, total_cell_ref, numeric_block_total = create_excel_table_for_block(output_ws, current_excel_row, 
                                                                        current_block_roman, current_block_title, processed_items_for_table)
                     recap_entries.append({'roman': current_block_roman, 'title': current_block_title, 
                                           'total_cell_ref': total_cell_ref, 'numeric_total': numeric_block_total})
                     current_excel_row = next_row # Update the current_excel_row for the next block

            current_block_roman = col_a_val
            current_block_title = col_b_val if col_b_val else f"Block {current_block_roman}"
            current_block_items = []
            print(f"Starting new block: {current_block_roman} - {current_block_title}")

        elif current_block_roman and col_b_val: 
            current_block_items.append((col_b_val, col_c_val, col_d_val_formula, col_e_val_pu))
        elif not current_block_roman and col_a_val and not is_roman:
            print(f"WARNING [parse_calcul_sheet]: Row {i_row+1} (ColA: '{col_a_val}', ColB: '{col_b_val}') before the first Roman block or unknown format, ignored.")


    if current_block_roman and current_block_items:
        print(f"\nProcessing last block (after loop): {current_block_roman} - {current_block_title} with {len(current_block_items)} items.")
        processed_items_for_table = []
        for item in current_block_items:
            desc, unit, formula, pu_raw = item
            qty_calculated = evaluate_formula(formula, qt_data, desc)
            unit_price = 0.0
            if pu_raw is not None:
                try:
                    if isinstance(pu_raw, str):
                        unit_price = float(pu_raw.replace(',','.'))
                    elif isinstance(pu_raw, (int, float)):
                        unit_price = float(pu_raw)
                except ValueError:
                    print(f"WARNING: Invalid unit price '{pu_raw}' for '{desc}'. Using 0.0.")
                    unit_price = 0.0
            
            processed_items_for_table.append([desc, unit, qty_calculated if qty_calculated is not None else 0.0, unit_price])
        
        if processed_items_for_table:
            next_row, total_cell_ref, numeric_block_total = create_excel_table_for_block(output_ws, current_excel_row, 
                                         current_block_roman, current_block_title, processed_items_for_table)
            recap_entries.append({'roman': current_block_roman, 'title': current_block_title, 
                                  'total_cell_ref': total_cell_ref, 'numeric_total': numeric_block_total})
            current_excel_row = next_row # Update the current_excel_row
    elif current_block_roman and not current_block_items:
         print(f"WARNING [parse_calcul_sheet]: The last block {current_block_roman} - {current_block_title} had no items to process.")
    
    return current_excel_row 

def process_menuiserie_block(open_data_list, output_ws, start_row, recap_entries):
    """
    Processes the 'open' data to format it for the 'Menuiserie' block (Bloc IV)
    and writes it to the output worksheet.
    
    Args:
        open_data_list (list): List of dictionaries, each representing an opening.
        output_ws (openpyxl.worksheet.worksheet.Worksheet): The output sheet to write results to.
        start_row (int): The row number from which to start writing this block.
        recap_entries (list): A list to append recap data (roman_numeral, title, total_cell_ref, numeric_total).
        
    Returns:
        int: The next available row number after processing the 'Menuiserie' block.
    """
    roman_numeral_main = "IV"
    header_title = "MENUISERIE"
    items_for_table = []

    for i, item_data in enumerate(open_data_list):
        designation = item_data.get('designation', '')
        largeur = item_data.get('l', 0.0)
        hauteur = item_data.get('h', 0.0)
        nombre = item_data.get('nombre', 0)
        type_ouverture = item_data.get('type', '')
        prix_unitaire = item_data.get('prix unitaire', 0.0)

        description = (f"Fourniture et pose de {designation}, {type_ouverture} "
                       f"({int(largeur*100)}X{int(hauteur*100)})")
        
        unit = "u" 
        qty = nombre 
        pu = prix_unitaire 
        
        items_for_table.append([description, unit, qty, pu])
    
    if items_for_table:
        next_row, total_cell_ref, numeric_block_total = create_excel_table_for_block(output_ws, start_row, 
                                                roman_numeral_main, header_title, items_for_table)
        recap_entries.append({'roman': roman_numeral_main, 'title': header_title, 
                              'total_cell_ref': total_cell_ref, 'numeric_total': numeric_block_total})
    else:
        print("WARNING: No valid items found for 'Menuiserie' block. Skipping table creation.")
        next_row = start_row 
        
    return next_row


def process_simple_block(data_list, output_ws, start_row, roman_numeral, header_title, item_start_num, recap_entries):
    """
    Processes 'simple' data (like Electricité/Plomberie with Designation, Unit, Number, Unit Price)
    and writes it to the output worksheet.
    
    Args:
        data_list (list): List of dictionaries from get_simple_block_data.
        output_ws (openpyxl.worksheet.worksheet.Worksheet): The output sheet.
        start_row (int): Starting row for this block.
        roman_numeral (str): Roman numeral for the block (e.g., "V").
        header_title (str): Title for the block (e.g., "ELECTRICITE").
        item_start_num (int): Starting number for items within the block (e.g., 1 for V.1).
        recap_entries (list): A list to append recap data (roman_numeral, title, total_cell_ref, numeric_total).
        
    Returns:
        int: The next available row number after processing this block.
    """
    items_for_table = []

    for i, item_data in enumerate(data_list):
        designation = item_data.get('designation', '')
        unit = item_data.get('unit', '')
        number = item_data.get('number', 0.0)
        unit_price = item_data.get('unit_price', 0.0)

        description = f"Fourniture et pose de {designation}"
        
        qty = number 
        pu = unit_price
        
        items_for_table.append([description, unit, qty, pu])
    
    if items_for_table:
        next_row, total_cell_ref, numeric_block_total = create_excel_table_for_block(output_ws, start_row, 
                                                roman_numeral, header_title, items_for_table)
        recap_entries.append({'roman': roman_numeral, 'title': header_title, 
                              'total_cell_ref': total_cell_ref, 'numeric_total': numeric_block_total})
    else:
        print(f"WARNING: No valid items found for '{header_title}' block. Skipping table creation.")
        next_row = start_row 
        
    return next_row

def process_formula_block(data_list, qt_data, output_ws, start_row, roman_numeral, header_title, item_start_num, recap_entries):
    """
    Processes 'formula' data (like Peinture/Revetement/Toiture with Description, Unit, Formula, P.U.)
    and writes it to the output worksheet, evaluating formulas using qt_data.
    
    Args:
        data_list (list): List of dictionaries from get_formula_block_data.
        qt_data (dict): The quantity data from the 'qt' sheet.
        output_ws (openpyxl.worksheet.worksheet.Worksheet): The output sheet.
        start_row (int): Starting row for this block.
        roman_numeral (str): Roman numeral for the block (e.g., "VII").
        header_title (str): Title for the block (e.g., "PEINTURE").
        item_start_num (int): Starting number for items within the block (e.g., 1 for VII.1).
        recap_entries (list): A list to append recap data (roman_numeral, title, total_cell_ref, numeric_total).
        
    Returns:
        int: The next available row number after processing this block.
    """
    items_for_table = []

    for i, item_data in enumerate(data_list):
        description = item_data.get('description', '')
        unit = item_data.get('unit', '')
        formula_or_qty = item_data.get('formula_or_qty', 0.0)
        pu_raw = item_data.get('pu', 0.0)

        qty_calculated = evaluate_formula(formula_or_qty, qt_data, description)
        
        unit_price = 0.0
        if pu_raw is not None:
            try:
                if isinstance(pu_raw, str):
                    unit_price = float(pu_raw.replace(',','.'))
                elif isinstance(pu_raw, (int, float)):
                    unit_price = float(pu_raw)
            except ValueError:
                print(f"WARNING: Invalid unit price '{pu_raw}' for '{description}'. Using 0.0.")
                unit_price = 0.0
        
        items_for_table.append([description, unit, qty_calculated if qty_calculated is not None else 0.0, unit_price])
    
    if items_for_table:
        next_row, total_cell_ref, numeric_block_total = create_excel_table_for_block(output_ws, start_row, 
                                                roman_numeral, header_title, items_for_table)
        recap_entries.append({'roman': roman_numeral, 'title': header_title, 
                              'total_cell_ref': total_cell_ref, 'numeric_total': numeric_block_total})
    else:
        print(f"WARNING: No valid items found for '{header_title}' block. Skipping table creation.")
        next_row = start_row 
        
    return next_row

def write_recap_block(output_ws, start_row, recap_entries):
    """
    Writes the recapitulative block to the Excel worksheet.
    
    Args:
        output_ws (openpyxl.worksheet.worksheet.Worksheet): The output sheet.
        start_row (int): The row number from which to start writing this block.
        recap_entries (list): A list of dictionaries, each containing
                              {'roman': str, 'title': str, 'total_cell_ref': str, 'numeric_total': float}.
    
    Returns:
        int: The next available row number after writing the recapitulative block.
    """
    thin_black_side = Side(style='thin', color='000000')
    black_border = Border(left=thin_black_side, right=thin_black_side, top=thin_black_side, bottom=thin_black_side)
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # Set column widths for recap table (A: Roman, B: Description, F: Montant)
    output_ws.column_dimensions['A'].width = 6
    output_ws.column_dimensions['B'].width = 60 # Description will be in B
    output_ws.column_dimensions['F'].width = 15 # Montant will be in F

    current_row = start_row
    
    # Title for the recapitulation
    output_ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    output_ws[f'A{current_row}'] = "--- RÉCAPITULATIF ---"
    output_ws[f'A{current_row}'].font = bold_font
    output_ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    output_ws[f'A{current_row}'].fill = header_fill
    current_row += 1

    # Write each block's summary
    recap_total_cell_refs = [] # To collect cell references for the grand total Excel formula
    grand_total_numeric_for_letter = 0.0 # To sum numeric totals for the letter conversion
    
    for entry in recap_entries:
        output_ws[f'A{current_row}'] = entry['roman']
        output_ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        output_ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=5)
        output_ws[f'B{current_row}'] = entry['title']
        output_ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        # Link to the total of the original block using Excel formula
        output_ws[f'F{current_row}'] = f"={entry['total_cell_ref']}"
        output_ws[f'F{current_row}'].number_format = '#,##0.00'
        output_ws[f'F{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
        
        recap_total_cell_refs.append(f'F{current_row}') # Add this cell to the list for grand total sum
        grand_total_numeric_for_letter += entry['numeric_total'] # Add numeric total for letter conversion
        current_row += 1
    
    # TOTAL GENERAL
    output_ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
    output_ws[f'A{current_row}'] = "TOTAL GÉNÉRAL HTVA"
    output_ws[f'A{current_row}'].font = bold_font
    output_ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
    
    grand_total_cell_ref = f'F{current_row}'
    if recap_total_cell_refs:
        grand_total_formula = f"=SUM({','.join(recap_total_cell_refs)})"
    else:
        grand_total_formula = 0
    output_ws[grand_total_cell_ref] = grand_total_formula
    output_ws[grand_total_cell_ref].font = bold_font
    output_ws[grand_total_cell_ref].number_format = '#,##0.00'
    output_ws[grand_total_cell_ref].alignment = Alignment(horizontal='right', vertical='center')
    current_row += 1

    # Apply borders to the recap table rows (except the final text line)
    for row_num in range(start_row, current_row):
        for col_idx in range(1, 7): # Columns A to F
            cell = output_ws.cell(row=row_num, column=col_idx)
            cell.border = black_border
    
    # Amount in letters
    text_total = conv_number_letter(round(grand_total_numeric_for_letter), devise=1, langue=0) # Round to nearest integer for currency in letters

    # Line for the amount in letters
    output_ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    output_ws[f'A{current_row}'] = f"Arrêter le présent devis estimatif à la somme de : {text_total}"
    output_ws[f'A{current_row}'].font = bold_font
    output_ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    output_ws[f'A{current_row}'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # Light grey fill for the text line
    current_row += 1

    return current_row # Return the next available row after the recap

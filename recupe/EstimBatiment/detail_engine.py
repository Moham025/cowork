# detail_engine.py
"""
Moteur de calcul du détail des matériaux.
Lit les feuilles calcul, qt, cf, ha pour produire :
  - nb sacs de ciment (TOTAL_ciment / 20)  [1 sac = 50 kg]
  - nb briques
  - nb hourdis
  - camions sable  (TOTAL_sable / 28)
  - camions granite (TOTAL_granite / 28)
  - nb planches 4.5m (TOTAL_planche / 4.5)
  - barres Ha6..Ha14 (TOTAL_ha / 12)
"""

import re, io
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── Colonnes matériaux dans la feuille Materiaux (0-based index) ─────────────
# A=0  B=1     C=2     D=3    E=4      F=5      G=6(eau) H=7    I=8
# J=9  K=10    L=11    M=12
MAT_COLS = {
    'ciment':  0,   # A
    'brique':  1,   # B
    'hourdi':  2,   # C
    'sable':   3,   # D
    'granite': 4,   # E
    'planche': 5,   # F
    'terre':   7,   # H  (G=eau ignoré)
    'ha6':     8,   # I
    'ha8':     9,   # J
    'ha10':    10,  # K
    'ha12':    11,  # L
    'ha14':    12,  # M
}
MAT_LABELS = ['ciment', 'brique', 'hourdi', 'sable', 'granite',
              'planche', 'terre', 'ha6', 'ha8', 'ha10', 'ha12', 'ha14']

ROMAN_HEADERS = {'I', 'II', 'III', 'IV', 'V', 'VI',
                 'VII', 'VIII', 'IX', 'X', 'XI', 'XII'}


# ═══════════════════════════════════════════════════════════════════════════════
#  Chargement des données cf et ha
# ═══════════════════════════════════════════════════════════════════════════════
def _load_cf(ws):
    """Retourne dict 'B7' -> valeur pour toutes cellules non vides de cf."""
    data = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                data[cell.coordinate.upper()] = cell.value
    return data


def _load_ha(ws):
    """Retourne dict 'F3' -> float pour toutes cellules numériques de ha."""
    data = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and isinstance(cell.value, (int, float)):
                data[cell.coordinate.upper()] = float(cell.value)
    return data


# ═══════════════════════════════════════════════════════════════════════════════
#  Évaluation des formules qt (colonne D du calcul)
# ═══════════════════════════════════════════════════════════════════════════════
def _eval_qt(formula, qt_data):
    """
    Résout les formules du type 'LONGRINE[ml]*0,02' en utilisant qt_data.
    Retourne un float.
    """
    if formula is None:
        return 0.0
    if isinstance(formula, (int, float)):
        return float(formula)

    s = str(formula).strip().replace(',', '.')

    def repl_qt(m):
        item = m.group(1).strip().lower()
        col  = m.group(2).strip().lower()
        v = (qt_data.get(item) or {}).get(col)
        return str(float(v)) if v is not None else '0.0'

    s = re.sub(r'([A-Za-z_0-9ÉÈÀÊÛÔÎÇ]+)\[([a-zA-Z0-9]+)\]', repl_qt, s)
    try:
        return float(eval(s, {"__builtins__": {}}, {}))
    except Exception:
        return 0.0


# ═══════════════════════════════════════════════════════════════════════════════
#  Évaluation des formules matériaux (colonnes G-S du calcul)
# ═══════════════════════════════════════════════════════════════════════════════
def _eval_mat(formula, row_d, d_vals, qt_data, cf_data, ha_data):
    """
    Évalue une formule de la colonne matériau.
    row_d   : valeur D (quantité) de la ligne courante
    d_vals  : dict {row_1based -> float} des valeurs D de toutes les lignes
    """
    if formula is None:
        return 0.0
    if isinstance(formula, (int, float)):
        return float(formula)

    s = str(formula).strip()

    # ── 'same' → même valeur que D ───────────────────────────────────────────
    if s.lower() == 'same':
        return row_d

    # ── 'xB2[cf]' → row_d * cf['B2'] ────────────────────────────────────────
    m = re.fullmatch(r'x([A-Za-z])(\d+)\[cf\]', s, re.IGNORECASE)
    if m:
        col, rn = m.group(1).upper(), int(m.group(2))
        cf_val = cf_data.get(f'{col}{rn}', cf_data.get(f'B{rn}', 0.0))
        return row_d * float(cf_val) if isinstance(cf_val, (int, float)) else 0.0

    # ── 'C8x0,6+C22x0,2[cf]' → d_vals[8]*0.6 + d_vals[22]*0.2 ──────────────
    if re.search(r'C\d+x', s, re.IGNORECASE):
        s2 = re.sub(r'\[cf\]$', '', s).replace(',', '.')
        s2 = re.sub(r'C(\d+)x', lambda mo: str(d_vals.get(int(mo.group(1)), 0.0)) + '*', s2,
                    flags=re.IGNORECASE)
        try:
            return float(eval(s2, {"__builtins__": {}}, {}))
        except Exception:
            return 0.0

    s = s.replace(',', '.')

    # ── Remplacer D{row}{calcul} ──────────────────────────────────────────────
    s = re.sub(r'[Dd](\d+)\{calcul\}',
               lambda mo: str(d_vals.get(int(mo.group(1)), 0.0)), s)

    # ── Remplacer {lettre}{row}{ha} (minuscule) : f3{ha} → ha['F3'] ──────────
    s = re.sub(r'([a-z])(\d+)\{ha\}',
               lambda mo: str(ha_data.get(f'{mo.group(1).upper()}{mo.group(2)}', 0.0)), s)

    # ── Remplacer {LETTRE}{row}{cf} : B7{cf}, D13{cf} ────────────────────────
    def repl_cf(mo):
        col, rn = mo.group(1).upper(), int(mo.group(2))
        v = cf_data.get(f'{col}{rn}', cf_data.get(f'B{rn}', 0.0))
        return str(float(v)) if isinstance(v, (int, float)) else '0.0'
    s = re.sub(r'([A-Za-z])(\d+)\{cf\}', repl_cf, s)

    # ── Remplacer ITEM[col] qt ────────────────────────────────────────────────
    def repl_qt(mo):
        item = mo.group(1).strip().lower()
        col  = mo.group(2).strip().lower()
        v = (qt_data.get(item) or {}).get(col)
        return str(float(v)) if v is not None else '0.0'
    s = re.sub(r'([A-Za-z_0-9ÉÈÀÊÛÔÎÇ]+)\[([a-zA-Z0-9]+)\]', repl_qt, s)

    try:
        return float(eval(s, {"__builtins__": {}}, {}))
    except Exception:
        return 0.0


# ═══════════════════════════════════════════════════════════════════════════════
#  Moteur principal
# ═══════════════════════════════════════════════════════════════════════════════
def compute_detail(file_bytes):
    """
    Prend les bytes du fichier Excel source, retourne (output_io, filename).
    """
    from data_reader import get_qt_data

    input_io = io.BytesIO(file_bytes)
    wb_f = openpyxl.load_workbook(input_io, data_only=False)
    input_io.seek(0)
    wb_v = openpyxl.load_workbook(input_io, data_only=True)

    for name in ('cf', 'ha', 'calcul', 'qt'):
        if name not in wb_v.sheetnames:
            raise ValueError(f"Feuille '{name}' absente du fichier.")

    qt_data  = get_qt_data(wb_v['qt'])
    cf_data  = _load_cf(wb_v['cf'])
    ha_data  = _load_ha(wb_v['ha'])
    calcul_f = wb_f['calcul']

    # ── Chargement feuille Materiaux (formules par numéro de ligne) ──────────
    # Supporte l'ancien format (colonnes G-S de calcul) et le nouveau (feuille Materiaux)
    mat_cells = {}   # {row_1based: {col_0based: formula_str}}
    if 'Materiaux' in wb_f.sheetnames:
        for row in wb_f['Materiaux'].iter_rows():
            for cell in row:
                if cell.value is not None:
                    mat_cells.setdefault(cell.row, {})[cell.column - 1] = cell.value
    use_mat_sheet = bool(mat_cells)

    # ── Passe 1 : calculer toutes les valeurs D (quantités) ──────────────────
    all_rows = list(calcul_f.iter_rows(min_row=1, values_only=True))
    d_vals = {}
    for i, row in enumerate(all_rows, start=1):
        if not row or all(v is None for v in row):
            continue
        cells = list(row) + [None] * max(0, 19 - len(row))
        a_str = str(cells[0]).strip().upper() if cells[0] else ''
        if a_str in ROMAN_HEADERS:
            continue
        d_formula = cells[3]
        if d_formula and cells[1]:
            d_vals[i] = _eval_qt(d_formula, qt_data)

    # ── Passe 2 : calculer les matériaux ligne par ligne ─────────────────────
    detail_rows = []
    current_roman = ''
    item_counters = {}

    for i, row in enumerate(all_rows, start=1):
        if not row or all(v is None for v in row):
            continue
        cells = list(row) + [None] * max(0, 19 - len(row))
        a     = cells[0]
        b     = cells[1]
        c     = cells[2]
        a_str = str(a).strip().upper() if a else ''

        # Section header
        if a_str in ROMAN_HEADERS:
            current_roman = a_str
            item_counters[current_roman] = 0
            detail_rows.append({
                'type': 'section',
                'num':  a_str,
                'desc': str(b) if b else '',
                **{k: None for k in MAT_LABELS},
            })
            continue

        # Ligne de totaux : détectée via feuille Materiaux (col A = TOTAL/xx)
        # ou via ancienne position col G de calcul
        if use_mat_sheet:
            mat_row = mat_cells.get(i, {})
            is_total = str(mat_row.get(0, '')).startswith('TOTAL')
        else:
            mat_row = {}
            is_total = str(cells[6] or '').startswith('TOTAL')
        if is_total:
            continue

        if not b:
            continue

        item_counters[current_roman] = item_counters.get(current_roman, 0) + 1
        num = f"{current_roman}.{item_counters[current_roman]}"

        row_d = d_vals.get(i, 0.0)

        mats = {}
        for mat, col_i in MAT_COLS.items():
            if use_mat_sheet:
                f_str = mat_row.get(col_i)
            else:
                f_str = cells[col_i] if col_i < len(cells) else None
            mats[mat] = _eval_mat(f_str, row_d, d_vals, qt_data, cf_data, ha_data)

        detail_rows.append({
            'type':  'item',
            'num':   num,
            'desc':  str(b) if b else '',
            'unite': str(c) if c else '',
            'qte':   row_d,
            **mats,
        })

    # ── Génération Excel ─────────────────────────────────────────────────────
    output_io = _write_excel(detail_rows)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return output_io, f"Detail_Materiaux_{ts}.xlsx"


# ═══════════════════════════════════════════════════════════════════════════════
#  Écriture du fichier Excel
# ═══════════════════════════════════════════════════════════════════════════════
def _write_excel(detail_rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Détail Matériaux"

    # ── Styles ────────────────────────────────────────────────────────────────
    thin  = Side(style='thin', color='000000')
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold  = Font(bold=True)
    white = Font(bold=True, color='FFFFFF')
    al_c  = Alignment(horizontal='center', vertical='center')
    al_r  = Alignment(horizontal='right',  vertical='center')
    al_l  = Alignment(horizontal='left',   vertical='center')
    al_cw = Alignment(horizontal='center', vertical='center', wrap_text=True)

    fill_hdr  = PatternFill('solid', fgColor='0D47A1')  # bleu entête tableau
    fill_sec  = PatternFill('solid', fgColor='1565C0')  # bleu section
    fill_tot  = PatternFill('solid', fgColor='E3F2FD')  # bleu clair total bloc
    fill_sum  = PatternFill('solid', fgColor='263238')  # sombre total général
    fill_res  = PatternFill('solid', fgColor='2E7D32')  # vert résumé
    fill_odd  = PatternFill('solid', fgColor='FAFAFA')
    fill_even = PatternFill('solid', fgColor='FFFFFF')

    num_fmt = '#,##0.00'

    # ── Colonnes ──────────────────────────────────────────────────────────────
    HEADERS = [
        ('N°',          6),
        ('Description', 42),
        ('Unité',        7),
        ('Quantité',    11),
        ('Ciment\n(sacs)', 10),
        ('Brique\n(nb)',   9),
        ('Hourdi\n(nb)',   9),
        ('Sable\n(m³)',    9),
        ('Granite\n(m³)',  9),
        ('Planche\n(m²)',  9),
        ('Terre\n(m³)',    9),
        ('Ha6\n(ml)',      9),
        ('Ha8\n(ml)',      9),
        ('Ha10\n(ml)',     9),
        ('Ha12\n(ml)',     9),
        ('Ha14\n(ml)',     9),
    ]
    N_COLS = len(HEADERS)

    for ci, (_, w) in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Titre principal ───────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N_COLS)
    ws['A1'] = "DÉTAIL DES MATÉRIAUX"
    ws['A1'].font      = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill      = fill_hdr
    ws['A1'].alignment = al_c
    ws['A1'].border    = brd

    # ── En-tête tableau ───────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 36
    for ci, (label, _) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=ci, value=label)
        cell.font      = white
        cell.fill      = fill_hdr
        cell.alignment = al_cw
        cell.border    = brd

    # ── Données ───────────────────────────────────────────────────────────────
    cur_row = 3
    totals  = {k: 0.0 for k in MAT_LABELS}
    sec_totals = {}
    sec_start_rows = {}
    current_sec = None
    odd = True

    def write_val(r, c, v, fmt=None):
        cell = ws.cell(row=r, column=c, value=v)
        if fmt:
            cell.number_format = fmt
        return cell

    for dr in detail_rows:
        if dr['type'] == 'section':
            # Sauvegarder la ligne de début pour les totaux de section
            current_sec = dr['num']
            sec_start_rows[current_sec] = cur_row
            sec_totals[current_sec] = {k: 0.0 for k in MAT_LABELS}

            ws.row_dimensions[cur_row].height = 18
            ws.merge_cells(start_row=cur_row, start_column=1,
                           end_row=cur_row, end_column=N_COLS)
            c = ws.cell(row=cur_row, column=1,
                        value=f"{dr['num']}  —  {dr['desc']}")
            c.font      = white
            c.fill      = fill_sec
            c.alignment = al_l
            c.border    = brd
            cur_row += 1

        else:
            ws.row_dimensions[cur_row].height = 16
            fill = fill_odd if odd else fill_even
            odd  = not odd

            vals = [
                dr.get('num', ''),
                dr.get('desc', ''),
                dr.get('unite', ''),
                dr.get('qte', 0.0),
            ] + [dr.get(k, 0.0) for k in MAT_LABELS]

            for ci, v in enumerate(vals, start=1):
                cell = ws.cell(row=cur_row, column=ci, value=v)
                cell.border    = brd
                cell.fill      = fill
                if ci == 2:
                    cell.alignment = al_l
                elif ci >= 4:
                    cell.alignment = al_r
                    if isinstance(v, float) and v != 0:
                        cell.number_format = '#,##0.##'
                else:
                    cell.alignment = al_c

            # Accumuler totaux
            for k in MAT_LABELS:
                v = dr.get(k, 0.0) or 0.0
                totals[k] += v
                if current_sec and current_sec in sec_totals:
                    sec_totals[current_sec][k] += v

            cur_row += 1

    # ── Ligne TOTAL GÉNÉRAL ───────────────────────────────────────────────────
    cur_row += 1
    ws.row_dimensions[cur_row].height = 20
    ws.merge_cells(start_row=cur_row, start_column=1,
                   end_row=cur_row, end_column=3)
    c = ws.cell(row=cur_row, column=1, value="TOTAL GÉNÉRAL")
    c.font = white; c.fill = fill_sum; c.alignment = al_l; c.border = brd

    for ci, k in enumerate(MAT_LABELS, start=5):
        cell = ws.cell(row=cur_row, column=ci, value=totals[k])
        cell.font          = white
        cell.fill          = fill_sum
        cell.border        = brd
        cell.alignment     = al_r
        cell.number_format = '#,##0.##'

    # col 4 (qte) vide
    ws.cell(row=cur_row, column=4).fill   = fill_sum
    ws.cell(row=cur_row, column=4).border = brd

    total_row = cur_row
    cur_row  += 2

    # ── RÉSUMÉ ────────────────────────────────────────────────────────────────
    ws.row_dimensions[cur_row].height = 22
    ws.merge_cells(start_row=cur_row, start_column=1,
                   end_row=cur_row, end_column=N_COLS)
    c = ws.cell(row=cur_row, column=1, value="RÉSUMÉ — APPROVISIONNEMENT")
    c.font = white; c.fill = fill_res; c.alignment = al_c; c.border = brd
    cur_row += 1

    resume_items = [
        ("Sacs de ciment  (50 kg/sac → tonnes)",
         totals['ciment'],      "sacs",    totals['ciment'] / 20,  "t"),
        ("Briques",
         totals['brique'],      "nb",      None,                   ""),
        ("Hourdis",
         totals['hourdi'],      "nb",      None,                   ""),
        ("Sable  (camions 28 m³)",
         totals['sable'],       "m³",      totals['sable'] / 28,   "camions"),
        ("Granite  (camions 28 m³)",
         totals['granite'],     "m³",      totals['granite'] / 28, "camions"),
        ("Planches",
         totals['planche'],     "nb",      None,                    "planches"),
        ("Terre  (camions 28 m³)",
         totals['terre'],       "m³",      totals['terre'] / 28,   "camions"),
        ("Fers Ha6  (barres 12 m)",
         totals['ha6'],         "ml",      totals['ha6'] / 12,     "barres"),
        ("Fers Ha8  (barres 12 m)",
         totals['ha8'],         "ml",      totals['ha8'] / 12,     "barres"),
        ("Fers Ha10 (barres 12 m)",
         totals['ha10'],        "ml",      totals['ha10'] / 12,    "barres"),
        ("Fers Ha12 (barres 12 m)",
         totals['ha12'],        "ml",      totals['ha12'] / 12,    "barres"),
        ("Fers Ha14 (barres 12 m)",
         totals['ha14'],        "ml",      totals['ha14'] / 12,    "barres"),
    ]

    fill_r_odd  = PatternFill('solid', fgColor='E8F5E9')
    fill_r_even = PatternFill('solid', fgColor='FFFFFF')

    for j, (label, total_val, unit_tot, result_val, unit_res) in enumerate(resume_items):
        ws.row_dimensions[cur_row].height = 18
        fill = fill_r_odd if j % 2 == 0 else fill_r_even

        ws.merge_cells(start_row=cur_row, start_column=1,
                       end_row=cur_row, end_column=5)
        c = ws.cell(row=cur_row, column=1, value=label)
        c.font = bold; c.fill = fill; c.alignment = al_l; c.border = brd

        # Total brut
        ws.merge_cells(start_row=cur_row, start_column=6,
                       end_row=cur_row, end_column=9)
        c2 = ws.cell(row=cur_row, column=6,
                     value=f"{total_val:,.2f} {unit_tot}".replace(',', ' ').replace('.', ','))
        c2.fill = fill; c2.alignment = al_r; c2.border = brd

        # Résultat (quantité commandable)
        ws.merge_cells(start_row=cur_row, start_column=10,
                       end_row=cur_row, end_column=N_COLS)
        if result_val is not None:
            import math
            qty = math.ceil(result_val)
            disp = f"→  {qty} {unit_res}"
        else:
            qty = math.ceil(total_val)
            disp = f"→  {qty} {unit_res}"
        c3 = ws.cell(row=cur_row, column=10, value=disp)
        c3.font = Font(bold=True, color='1B5E20')
        c3.fill = fill; c3.alignment = al_l; c3.border = brd

        cur_row += 1

    # ── Freeze panes ─────────────────────────────────────────────────────────
    ws.freeze_panes = 'C3'

    output_io = io.BytesIO()
    wb.save(output_io)
    output_io.seek(0)
    return output_io
